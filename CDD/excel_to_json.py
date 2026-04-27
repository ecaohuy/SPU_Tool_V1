import argparse
import json
import re
from pathlib import Path

import openpyxl

SITE_COL = "F"
SECTOR_COLS = {
    "sector_1": "CH",
    "sector_2": "CI",
    "sector_3": "CJ",
    "sector_4": "CK",
}
SECTOR_5G_COLS = {
    "sector_1": "CL",
    "sector_2": "CM",
    "sector_3": "CN",
    "sector_4": "CO",
}
DATA_START_ROW = 4

TOKEN_RE = re.compile(r"^\s*(?P<band>[^-\s]+)\s*-\s*(?P<config>.+?)\s*$")
TOKEN_5G_RE = re.compile(
    r"^\s*(?P<band>5G)\s+(?P<bw>\d+(?:\.\d+)?)\s*(?P<unit>MHz)?\s*$",
    re.IGNORECASE,
)


def parse_sector(cell_value):
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    out = []
    for tok in (t.strip() for t in text.split(";")):
        if not tok:
            continue
        m = TOKEN_RE.match(tok)
        if m:
            out.append({"band": m.group("band"), "config": m.group("config")})
        else:
            out.append({"band": tok, "config": ""})
    return out


def parse_sector_5g(cell_value):
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    out = []
    for tok in (t.strip() for t in text.split(";")):
        if not tok:
            continue
        m = TOKEN_5G_RE.match(tok)
        if m:
            unit = m.group("unit") or "MHz"
            out.append({"band": "5G", "config": f"{m.group('bw')} {unit}"})
        else:
            out.append({"band": tok, "config": ""})
    return out


def convert(xlsx_path: Path, sheet: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    sites = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        site = ws[f"{SITE_COL}{row}"].value
        if site is None or not str(site).strip():
            continue
        sectors_5g = {}
        for key, col in SECTOR_5G_COLS.items():
            parsed = parse_sector_5g(ws[f"{col}{row}"].value)
            if parsed:
                sectors_5g[key] = parsed
        # 4G TDD shares the AAU with 5G -> route TDD entries into sectors_5G.
        sectors_4g = {}
        for key, col in SECTOR_COLS.items():
            parsed = parse_sector(ws[f"{col}{row}"].value)
            if not parsed:
                continue
            fdd = [e for e in parsed if str(e.get("config", "")).strip().upper() != "TDD"]
            tdd = [e for e in parsed if str(e.get("config", "")).strip().upper() == "TDD"]
            if fdd:
                sectors_4g[key] = fdd
            if tdd:
                sectors_5g.setdefault(key, []).extend(tdd)
        site_obj = {"site_name": str(site).strip(), "sectors_4G": sectors_4g}
        if sectors_5g:
            site_obj["sectors_5G"] = sectors_5g
        sites.append(site_obj)
    return sites


def default_output_path(xlsx_path: Path) -> Path:
    return xlsx_path.resolve().parents[1] / "Output" / f"{xlsx_path.stem}.json"


def main():
    ap = argparse.ArgumentParser(description="Convert Call-off Excel to site/sector JSON.")
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--out", type=Path, default=None,
                    help="Defaults to <repo>/Output/<input-stem>.json")
    args = ap.parse_args()

    data = convert(args.xlsx, args.sheet)
    out = args.out or default_output_path(args.xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(data)} sites -> {out}")


if __name__ == "__main__":
    main()
