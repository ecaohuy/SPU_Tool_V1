import argparse
import json
import re
from pathlib import Path

import openpyxl

CELLNAME_COL = 4
MIMO_COL = 6
RRU_COL = 15
RRUNAME_COL = 16
RRUPORT_COL = 17
RIPORT_BB_COL = 18
DATA_START_ROW = 5
DEFAULT_SHEET = "Radio 4G"

RADIO_5G_SHEET = "Radio 5G"
CELLNAME_5G_COL = 3
RRU_5G_COL = 17
RRUNAME_5G_COL = 18
RRUPORT_5G_COL = 19
RIPORT_BB_5G_COL = 20
DATA_START_ROW_5G = 5

IP_SHEET = "IP"
IP_NE_COL = 1
IP_BASEBAND_COL = 15
IP_DATA_START_ROW = 5
NE_4G_COL = 1
CELLTYPE_4G_COL = 5
NE_5G_COL = 1
NRCELL_5G_COL = 3

BASEBAND_MIXED = "10G;e2q_3:3NR+6TDL;e2q_6:24FDL"
BASEBAND_FDD_ONLY = "10G;e2q_3:6NR;e2q_6:24FDL"

CELL_RE = re.compile(
    r"^e(?P<site>[A-Z0-9]{7})NR(?P<prefix>0\d1)B(?P<band>\d{2})(?P<suffix>\d)$"
)
CELL_5G_RE = re.compile(
    r"^g(?P<site>[A-Z0-9]+)_(?P<prefix>\d{2})n(?P<band>\d+)(?P<suffix>\d+)$"
)
RRU_5G_MAP = {
    "10": ("AAU_1", "VBP_1_3&OF1"),
    "20": ("AAU_2", "VBP_1_3&OF2"),
    "30": ("AAU_3", "VBP_1_3&OF3"),
}
BAND_MAP = {"28": "700", "03": "1800", "01": "2100", "41": "2600"}

RRUNAME_FIXED = {
    "700": "R9264_M7290N(ACA)",
    "2100": "R9264S_M1821(AAA)",
    "2600": "A9651A_S26",
}

RRU_RIPORT_MAP = {
    "2600": {
        "011": ("AAU_1", "VBP_1_3&OF1"),
        "021": ("AAU_2", "VBP_1_3&OF2"),
        "031": ("AAU_3", "VBP_1_3&OF3"),
    },
    "700": {
        "011": ("RRU_4", "VBP_1_6&OF4"),
        "021": ("RRU_5", "VBP_1_6&OF5"),
        "031": ("RRU_6", "VBP_1_6&OF6"),
    },
    "1800": {
        "011": ("RRU_1", "VBP_1_6&OF1"),
        "021": ("RRU_2", "VBP_1_6&OF2"),
        "031": ("RRU_3", "VBP_1_6&OF3"),
    },
    "2100": {
        "011": ("RRU_1", "VBP_1_6&OF1"),
        "021": ("RRU_2", "VBP_1_6&OF2"),
        "031": ("RRU_3", "VBP_1_6&OF3"),
    },
}

RRUPORT_BY_MIMO = {
    "4T4R": "1-4",
    "2T2R": "1-2",
    "64T64R": "1-64",
}


def normalize_site(cell_site: str) -> str:
    return cell_site[:2] + "U" + cell_site[3:]


def build_lookup(json_data):
    lookup = {}
    for site in json_data:
        name = site["site_name"]
        for sector_key, entries in site.get("sectors_4G", {}).items():
            for entry in entries:
                lookup[(name, sector_key, str(entry["band"]))] = str(entry["config"])
    return lookup


def mimo_for(band, site, sector_key, lookup):
    if band == "2600":
        return "64T64R", None
    if band not in {"700", "1800", "2100"}:
        return None, f"band_{band}_no_mimo_rule"
    config = lookup.get((site, sector_key, band))
    if config is None:
        return None, f"no_json_entry_{site}/{sector_key}/{band}"
    cfg = config.strip().upper()
    if "4T" in cfg:
        return "4T4R", None
    if "2T" in cfg:
        return "2T2R", None
    return None, f"config_{config}_unmapped"


def collect_ne_radio_info(wb):
    info = {}
    if DEFAULT_SHEET in wb.sheetnames:
        ws = wb[DEFAULT_SHEET]
        for r in range(DATA_START_ROW, ws.max_row + 1):
            ne = ws.cell(row=r, column=NE_4G_COL).value
            if not ne:
                continue
            entry = info.setdefault(str(ne).strip(), {"celltypes_4g": set(), "has_5g": False})
            ct = ws.cell(row=r, column=CELLTYPE_4G_COL).value
            if ct:
                entry["celltypes_4g"].add(str(ct).strip().upper())
    if RADIO_5G_SHEET in wb.sheetnames:
        ws = wb[RADIO_5G_SHEET]
        for r in range(DATA_START_ROW_5G, ws.max_row + 1):
            ne = ws.cell(row=r, column=NE_5G_COL).value
            if not ne:
                continue
            entry = info.setdefault(str(ne).strip(), {"celltypes_4g": set(), "has_5g": False})
            if ws.cell(row=r, column=NRCELL_5G_COL).value:
                entry["has_5g"] = True
    return info


def baseband_config_for(ne_info):
    if not ne_info:
        return None, "ne_not_in_radio_sheets"
    if not ne_info["has_5g"]:
        return None, "no_5g_nRCell"
    cts = ne_info["celltypes_4g"]
    if "TDD" in cts and "FDD" in cts:
        return BASEBAND_MIXED, None
    if cts == {"FDD"}:
        return BASEBAND_FDD_ONLY, None
    return None, f"4g_celltypes_{sorted(cts)}_no_rule"


def rruname_for(band, site, sector_key, lookup):
    if band in RRUNAME_FIXED:
        return RRUNAME_FIXED[band], None
    if band == "1800":
        if (site, sector_key, "2100") in lookup:
            return "R9264S_M1821(AAA)", None
        return None, f"1800_no_2100_in_{site}/{sector_key}"
    return None, f"band_{band}_no_rruname_rule"


def fill_workbook(xlsx_path: Path, json_path: Path, out_path: Path | None = None,
                  sheet: str = DEFAULT_SHEET):
    """Apply MIMO/RRU/RRUname/rruPort/RiPort fills to the Radio 4G + Radio 5G sheets.

    Returns (counts_4g, counts_5g, skipped, out_path).
    """
    json_data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    lookup = build_lookup(json_data)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet]

    counts = {"mimo": 0, "rruname": 0, "rruport": 0, "rru": 0, "riport_bb": 0}
    skipped = []
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=CELLNAME_COL).value
        if not cell_value:
            continue
        m = CELL_RE.match(str(cell_value).strip())
        if not m:
            skipped.append((row_idx, cell_value, "cellname_pattern_mismatch"))
            continue
        band_code = m.group("band")
        prefix = m.group("prefix")
        if band_code not in BAND_MAP:
            skipped.append((row_idx, cell_value, f"band_B{band_code}_not_in_rule"))
            continue
        site = normalize_site(m.group("site"))
        sector_key = f"sector_{int(prefix[1])}"
        band = BAND_MAP[band_code]

        mimo, reason = mimo_for(band, site, sector_key, lookup)
        if mimo:
            ws.cell(row=row_idx, column=MIMO_COL).value = mimo
            counts["mimo"] += 1
            port = RRUPORT_BY_MIMO.get(mimo)
            if port:
                ws.cell(row=row_idx, column=RRUPORT_COL).value = port
                counts["rruport"] += 1
            else:
                skipped.append((row_idx, cell_value, f"mimo_{mimo}_no_rruport_rule"))
        else:
            skipped.append((row_idx, cell_value, f"mimo: {reason}"))

        rruname, reason = rruname_for(band, site, sector_key, lookup)
        if rruname:
            ws.cell(row=row_idx, column=RRUNAME_COL).value = rruname
            counts["rruname"] += 1
        else:
            skipped.append((row_idx, cell_value, f"rruname: {reason}"))

        rru_pair = RRU_RIPORT_MAP.get(band, {}).get(prefix)
        if rru_pair:
            rru, riport_bb = rru_pair
            ws.cell(row=row_idx, column=RRU_COL).value = rru
            ws.cell(row=row_idx, column=RIPORT_BB_COL).value = riport_bb
            counts["rru"] += 1
            counts["riport_bb"] += 1
        else:
            skipped.append((row_idx, cell_value, f"rru/riport: band_{band}_prefix_{prefix}_no_rule"))

    counts_5g = {"rru": 0, "rruname": 0, "rruport": 0, "riport_bb": 0}
    if RADIO_5G_SHEET in wb.sheetnames:
        ws5 = wb[RADIO_5G_SHEET]
        for row_idx in range(DATA_START_ROW_5G, ws5.max_row + 1):
            cell_value = ws5.cell(row=row_idx, column=CELLNAME_5G_COL).value
            if not cell_value:
                continue
            m = CELL_5G_RE.match(str(cell_value).strip())
            if not m:
                skipped.append((row_idx, cell_value, "5G cellname_pattern_mismatch"))
                continue
            prefix = m.group("prefix")
            rru_pair = RRU_5G_MAP.get(prefix)
            if not rru_pair:
                skipped.append((row_idx, cell_value, f"5G prefix_{prefix}_no_rule"))
                continue
            rru, riport_bb = rru_pair
            ws5.cell(row=row_idx, column=RRU_5G_COL).value = rru
            ws5.cell(row=row_idx, column=RRUNAME_5G_COL).value = "A9651A_S26"
            ws5.cell(row=row_idx, column=RRUPORT_5G_COL).value = "1-64"
            ws5.cell(row=row_idx, column=RIPORT_BB_5G_COL).value = riport_bb
            counts_5g["rru"] += 1
            counts_5g["rruname"] += 1
            counts_5g["rruport"] += 1
            counts_5g["riport_bb"] += 1

    counts["ip_baseband"] = 0
    if IP_SHEET in wb.sheetnames:
        ne_info = collect_ne_radio_info(wb)
        ws_ip = wb[IP_SHEET]
        for r in range(IP_DATA_START_ROW, ws_ip.max_row + 1):
            ne = ws_ip.cell(row=r, column=IP_NE_COL).value
            if not ne:
                continue
            bb, reason = baseband_config_for(ne_info.get(str(ne).strip()))
            if bb:
                ws_ip.cell(row=r, column=IP_BASEBAND_COL).value = bb
                counts["ip_baseband"] += 1
            else:
                skipped.append((r, ne, f"IP baseband: {reason}"))

    out_path = Path(out_path) if out_path else xlsx_path.with_name(xlsx_path.stem + "_filled.xlsx")
    wb.save(out_path)
    return counts, counts_5g, skipped, out_path


def main():
    ap = argparse.ArgumentParser(description="Fill Radio 4G/5G columns from Call-off JSON.")
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("json_path", type=Path)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--out", type=Path, default=None,
                    help="Defaults to <input>_filled.xlsx next to input.")
    args = ap.parse_args()

    counts, counts_5g, skipped, out_path = fill_workbook(args.xlsx, args.json_path, args.out, args.sheet)
    print(
        f"4G filled MIMO:{counts['mimo']} rruPort:{counts['rruport']} "
        f"RRU:{counts['rru']} RRUname:{counts['rruname']} "
        f"RiPortBB:{counts['riport_bb']}"
    )
    print(f"IP Baseband config filled: {counts.get('ip_baseband', 0)}")
    print(
        f"5G filled RRU:{counts_5g['rru']} RRUname:{counts_5g['rruname']} "
        f"rruPort:{counts_5g['rruport']} RiPortBB:{counts_5g['riport_bb']}"
    )
    print(f"-> {out_path}")
    if skipped:
        print(f"Skipped {len(skipped)} entries:")
        for row_idx, cell, reason in skipped:
            print(f"  r{row_idx}: {cell!r} -> {reason}")


if __name__ == "__main__":
    main()
