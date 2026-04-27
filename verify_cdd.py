"""Verify a CDD xlsx workbook against the rules in CLAUDE.md.

Usage:
    uv run --with openpyxl python verify_cdd.py <path-to-xlsx>
    uv run --with openpyxl python verify_cdd.py --fix <path-to-xlsx>

Without --fix: validates and writes <input>.findings.json/.csv. Exit 0 if clean.
With --fix:   applies deterministic corrections (NE_Name from IP, eNBId/gNBId
from IP, 4G TDD/FDD CellType + RRUname + RRU/RiPort prefix swap, 5G constants)
into a copy named <input>_fixed.xlsx and writes a residual findings report
beside it. Non-deterministic rules (e.g. 5G bSChannelBwDL/arfcnDL inherited
from 4G TDD) are reported, never auto-changed.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl

DATA_START_ROW = 5  # rows 1-4 are header/meta in every Radio sheet

# Radio 4G column letters -> 0-indexed positions
R4G = {
    "NE_Name": 0,        # A
    "eNBId": 1,          # B
    "CellName": 3,       # D  (CellName / userLabel)
    "CellType": 4,       # E
    "cpSpeRefSigPwr": 8, # I
    "maxCPTransPwr": 9,  # J
    "arfcndl": 11,       # L
    "arfcnul": 12,       # M
    "dlChannelBandwidth": 13,  # N
    "RRU": 14,           # O
    "RRUname": 15,       # P
    "rruPort": 16,       # Q  (TxRxChannelNo)
    "RiPort": 17,        # R  (RiPort Baseband)
}

# Radio 5G column letters -> 0-indexed positions
R5G = {
    "NE_Name": 0,        # A
    "gNBId": 1,          # B
    "arfcnDL": 11,       # L
    "arfcnUL": 12,       # M
    "bSChannelBwDL": 13, # N
    "configuredMaxTxPower": 14,  # O
    "CellType": 15,      # P
    "RRU": 16,           # Q
    "RRUname": 17,       # R
    "rruPort": 18,       # S
    "RiPort": 19,        # T
}


@dataclass
class Finding:
    sheet: str
    row: int
    column: str
    ne_name: str
    rule_id: str
    severity: str
    actual: object
    expected: object
    message: str


# ------------------------- parsers -------------------------

@dataclass
class IpRow:
    row: int
    ne_name: str
    enb_id: object
    gnb_id: object


def parse_ip(ws) -> dict[str, IpRow]:
    """Index IP sheet by NE_Name. NE_Name is column A, eNBId E (idx 4), gNBId J (idx 9)."""
    out: dict[str, IpRow] = {}
    for i, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        ne = row[0]
        if not ne:
            continue
        out[str(ne)] = IpRow(row=i, ne_name=str(ne), enb_id=row[4], gnb_id=row[9])
    return out


@dataclass
class RadioRow:
    sheet: str
    row: int
    values: tuple
    cols: dict[str, int]

    def get(self, key: str):
        idx = self.cols[key]
        return self.values[idx] if idx < len(self.values) else None


def parse_radio(ws, sheet: str, cols: dict[str, int]) -> list[RadioRow]:
    rows: list[RadioRow] = []
    for i, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        if not row or not row[0]:
            continue
        rows.append(RadioRow(sheet=sheet, row=i, values=row, cols=cols))
    return rows


# ------------------------- NE_Name mapping -------------------------

def expected_4g_ne(ip_ne: str) -> str:
    """Per CLAUDE.md, Radio 4G NE_Name must equal the IP NE_Name."""
    return ip_ne


def ip_ne_for_4g(radio_ne: str) -> str | None:
    """Heuristic: collapse legacy 'e<X>NR0' wording to its IP 'g<X>' form so
    typo'd rows can still be matched back to an IP entry."""
    if not radio_ne:
        return None
    base = radio_ne
    if base.endswith("NR0"):
        base = base[:-3]
    if base.startswith("e"):
        base = "g" + base[1:]
    return base


def best_ip_match(token: str, ip: dict[str, IpRow]) -> str | None:
    """Best-effort match for typo'd Radio NE_Name -> IP NE_Name (longest common prefix)."""
    if not token:
        return None
    best, best_score = None, 0
    for k in ip:
        n = min(len(k), len(token))
        score = sum(1 for j in range(n) if k[j] == token[j])
        if score > best_score:
            best, best_score = k, score
    return best


# ------------------------- rules -------------------------

def r_ne_name_4g(ip: dict[str, IpRow], rows4g: list[RadioRow]) -> list[Finding]:
    findings: list[Finding] = []
    valid = {expected_4g_ne(k) for k in ip}
    for r in rows4g:
        ne = str(r.get("NE_Name") or "")
        if ne in valid:
            continue
        guess = best_ip_match(ip_ne_for_4g(ne) or "", ip)
        expected = expected_4g_ne(guess) if guess else None
        findings.append(Finding(
            sheet=r.sheet, row=r.row, column="NE_Name", ne_name=ne,
            rule_id="r_ne_name_4g", severity="error",
            actual=ne, expected=expected,
            message=f"Radio 4G NE_Name '{ne}' does not match any IP NE_Name (expected '{expected}')",
        ))
    return findings


def r_ne_name_5g(ip: dict[str, IpRow], rows5g: list[RadioRow]) -> list[Finding]:
    findings: list[Finding] = []
    for r in rows5g:
        ne = str(r.get("NE_Name") or "")
        if ne in ip:
            continue
        guess = best_ip_match(ne, ip)
        findings.append(Finding(
            sheet=r.sheet, row=r.row, column="NE_Name", ne_name=ne,
            rule_id="r_ne_name_5g", severity="error",
            actual=ne, expected=guess,
            message=f"Radio 5G NE_Name '{ne}' is not in IP sheet (expected '{guess}')",
        ))
    return findings


def r_enb_id(ip: dict[str, IpRow], rows4g: list[RadioRow]) -> list[Finding]:
    findings: list[Finding] = []
    for r in rows4g:
        ne = str(r.get("NE_Name") or "")
        ip_ne = ip_ne_for_4g(ne) or ""
        ip_row = ip.get(ip_ne) or ip.get(best_ip_match(ip_ne, ip) or "")
        if not ip_row:
            continue  # NE_Name issue already flagged by r_ne_name_4g
        actual = r.get("eNBId")
        if actual != ip_row.enb_id:
            findings.append(Finding(
                sheet=r.sheet, row=r.row, column="eNBId", ne_name=ne,
                rule_id="r_enb_id", severity="error",
                actual=actual, expected=ip_row.enb_id,
                message=f"eNBId {actual!r} does not match IP value {ip_row.enb_id!r} for {ip_row.ne_name}",
            ))
    return findings


def r_gnb_id(ip: dict[str, IpRow], rows5g: list[RadioRow]) -> list[Finding]:
    findings: list[Finding] = []
    for r in rows5g:
        ne = str(r.get("NE_Name") or "")
        ip_row = ip.get(ne) or ip.get(best_ip_match(ne, ip) or "")
        if not ip_row:
            continue
        actual = r.get("gNBId")
        if actual != ip_row.gnb_id:
            findings.append(Finding(
                sheet=r.sheet, row=r.row, column="gNBId", ne_name=ne,
                rule_id="r_gnb_id", severity="error",
                actual=actual, expected=ip_row.gnb_id,
                message=f"gNBId {actual!r} does not match IP value {ip_row.gnb_id!r} for {ip_row.ne_name}",
            ))
    return findings


def r_4g_tdd_fdd(rows4g: list[RadioRow]) -> list[Finding]:
    findings: list[Finding] = []
    for r in rows4g:
        ne = str(r.get("NE_Name") or "")
        dl, ul = r.get("arfcndl"), r.get("arfcnul")
        if dl is None or ul is None:
            continue
        is_tdd = dl == ul
        ct = str(r.get("CellType") or "")
        rru = str(r.get("RRU") or "")
        rruname = str(r.get("RRUname") or "")
        rip = str(r.get("RiPort") or "")

        if is_tdd:
            checks = [
                ("CellType", ct == "TDD", ct, "TDD"),
                ("RRU", rru.startswith("AAU_"), rru, "AAU_…"),
                ("RRUname", rruname == "A9651A_S26", rruname, "A9651A_S26"),
                ("RiPort", rip.startswith("VBP_1_3"), rip, "VBP_1_3…"),
            ]
            tag = "TDD"
        else:
            checks = [
                ("CellType", ct == "FDD", ct, "FDD"),
                ("RRU", rru.startswith("RRU_"), rru, "RRU_…"),
                ("RRUname", rruname.startswith("R"), rruname, "R…"),
                ("RiPort", rip.startswith("VBP_1_6"), rip, "VBP_1_6…"),
            ]
            tag = "FDD"
        for col, ok, actual, expected in checks:
            if not ok:
                findings.append(Finding(
                    sheet=r.sheet, row=r.row, column=col, ne_name=ne,
                    rule_id="r_4g_tdd_fdd", severity="error",
                    actual=actual, expected=expected,
                    message=f"Radio 4G {tag} row (arfcndl={dl}, arfcnul={ul}) expects {col} {expected!r}, got {actual!r}",
                ))
    return findings


_B41_SECTOR_RE = re.compile(r"(\d{3})B41")
_AAU_RE = re.compile(r"AAU[-_]?(\d+)")


def sector_from_cellname_for_band(cellname: str, band: str) -> int | None:
    """Sector index from a CellName like 'eCMT8905NR011B28': 3 digits before
    `band`, take the middle digit. '011' -> 1, '021' -> 2, '031' -> 3."""
    if not cellname or not band:
        return None
    m = re.search(rf"(\d{{3}}){re.escape(band)}", cellname)
    if not m:
        return None
    middle = m.group(1)[1]
    return int(middle) if middle.isdigit() else None


# (RRUname, band marker in CellName, RRU index offset)
# Sector (1/2/3) + offset = RRU_<n> and VBP_1_6&OF<n>.
#  - R9264_M7290N(ACA)   on B28: sectors 1/2/3 -> RRU_4/5/6 (offset +3)
#  - R9264S_M1821(AAA)   on B03: sectors 1/2/3 -> RRU_1/2/3 (offset 0)
#  - R9264S_M1821(AAA)   on B01: sectors 1/2/3 -> RRU_1/2/3 (offset 0)
_RRUNAME_BAND_RULES: list[tuple[str, str, int]] = [
    ("R9264_M7290N(ACA)", "B28", 3),
    ("R9264S_M1821(AAA)", "B03", 0),
    ("R9264S_M1821(AAA)", "B01", 0),
]


def sector_from_rru(rru) -> int | None:
    """Sector index from an RRU value like 'AAU_1', 'AAU-2'."""
    if rru is None:
        return None
    m = _AAU_RE.search(str(rru))
    return int(m.group(1)) if m else None


def normalize_to_ip_ne(ne: str, ip: dict, *, from_4g: bool) -> str | None:
    """Map a Radio NE_Name (possibly typo'd or in legacy 4G 'e<X>NR0' form) to
    its IP NE_Name. Returns None if no plausible match exists."""
    if not ne:
        return None
    candidate = ip_ne_for_4g(ne) if from_4g else ne
    if candidate in ip:
        return candidate
    return best_ip_match(candidate or ne, ip)


def sector_from_cellname(cellname: str) -> int | None:
    """In a CellName like 'eCMT8905NR011B411', the 3 chars before 'B41' encode
    the sector: '011' -> 1, '021' -> 2, '031' -> 3 (middle digit). Returns
    None when the pattern is not present."""
    if not cellname:
        return None
    m = _B41_SECTOR_RE.search(cellname)
    if not m:
        return None
    middle = m.group(1)[1]
    if not middle.isdigit():
        return None
    return int(middle)


def r_4g_rru_aau_from_cellname(rows4g: list[RadioRow]) -> list[Finding]:
    """Radio 4G B41 rows: RRU must be 'AAU_<sector>' where sector is derived
    from the CellName (3 chars before 'B41', take the middle digit)."""
    findings: list[Finding] = []
    for r in rows4g:
        cellname = str(r.get("CellName") or "")
        sector = sector_from_cellname(cellname)
        if sector is None:
            continue
        expected = f"AAU_{sector}"
        actual = str(r.get("RRU") or "")
        if actual != expected:
            findings.append(Finding(
                sheet=r.sheet, row=r.row, column="RRU",
                ne_name=str(r.get("NE_Name") or ""),
                rule_id="r_4g_rru_aau_from_cellname", severity="error",
                actual=actual, expected=expected,
                message=f"RRU should be {expected!r} (sector from CellName {cellname!r}), got {actual!r}",
            ))
    return findings


_R4G_TDD_POWER_TARGETS = {
    "cpSpeRefSigPwr": 9.2,
    "maxCPTransPwr": 40,
}


def r_4g_tdd_power(rows4g: list[RadioRow]) -> list[Finding]:
    """Radio 4G TDD rows (arfcndl == arfcnul) must have cpSpeRefSigPwr=9.2 and
    maxCPTransPwr=40. Frequency-based, like the other TDD rules — so it still
    fires when CellType is mislabeled and only fixed by r_4g_tdd_fdd."""
    findings: list[Finding] = []
    for r in rows4g:
        dl, ul = r.get("arfcndl"), r.get("arfcnul")
        if dl is None or ul is None or dl != ul:
            continue
        ne = str(r.get("NE_Name") or "")
        for col, expected in _R4G_TDD_POWER_TARGETS.items():
            actual = r.get(col)
            try:
                if actual is not None and float(actual) == float(expected):
                    continue
            except (TypeError, ValueError):
                pass
            findings.append(Finding(
                sheet=r.sheet, row=r.row, column=col, ne_name=ne,
                rule_id="r_4g_tdd_power", severity="error",
                actual=actual, expected=expected,
                message=f"Radio 4G TDD row expects {col}={expected}, got {actual!r}",
            ))
    return findings


def r_4g_rruname_riport_rru_mapping(rows4g: list[RadioRow]) -> list[Finding]:
    """Specific RRUname + CellName-band combinations dictate exact RiPort and
    RRU values per sector (see _RRUNAME_BAND_RULES)."""
    findings: list[Finding] = []
    for r in rows4g:
        rruname = str(r.get("RRUname") or "")
        cellname = str(r.get("CellName") or "")
        for cfg_name, band, offset in _RRUNAME_BAND_RULES:
            if rruname != cfg_name:
                continue
            sector = sector_from_cellname_for_band(cellname, band)
            if sector is None:
                continue
            idx = sector + offset
            expected_riport = f"VBP_1_6&OF{idx}"
            expected_rru = f"RRU_{idx}"
            actual_riport = str(r.get("RiPort") or "")
            actual_rru = str(r.get("RRU") or "")
            ne = str(r.get("NE_Name") or "")
            if actual_riport != expected_riport:
                findings.append(Finding(
                    sheet=r.sheet, row=r.row, column="RiPort", ne_name=ne,
                    rule_id="r_4g_rruname_band_mapping", severity="error",
                    actual=actual_riport, expected=expected_riport,
                    message=(f"RRUname {rruname!r} on {band} sector {sector} "
                             f"expects RiPort {expected_riport!r}, got {actual_riport!r}"),
                ))
            if actual_rru != expected_rru:
                findings.append(Finding(
                    sheet=r.sheet, row=r.row, column="RRU", ne_name=ne,
                    rule_id="r_4g_rruname_band_mapping", severity="error",
                    actual=actual_rru, expected=expected_rru,
                    message=(f"RRUname {rruname!r} on {band} sector {sector} "
                             f"expects RRU {expected_rru!r}, got {actual_rru!r}"),
                ))
            break
    return findings


def r_4g_rruport_tdd(rows4g: list[RadioRow]) -> list[Finding]:
    """Radio 4G TDD rows (arfcndl == arfcnul): rruPort must be '1-64'."""
    findings: list[Finding] = []
    for r in rows4g:
        dl, ul = r.get("arfcndl"), r.get("arfcnul")
        if dl is None or ul is None or dl != ul:
            continue
        actual = str(r.get("rruPort") or "")
        if actual != "1-64":
            findings.append(Finding(
                sheet=r.sheet, row=r.row, column="rruPort",
                ne_name=str(r.get("NE_Name") or ""),
                rule_id="r_4g_rruport_tdd", severity="error",
                actual=actual, expected="1-64",
                message=f"Radio 4G TDD row expects rruPort '1-64', got {actual!r}",
            ))
    return findings


def r_5g_tdd_consistency(rows5g: list[RadioRow]) -> list[Finding]:
    findings: list[Finding] = []
    for r in rows5g:
        ne = str(r.get("NE_Name") or "")
        ct = str(r.get("CellType") or "")
        rru = str(r.get("RRU") or "")
        rruname = str(r.get("RRUname") or "")
        rrup = str(r.get("rruPort") or "")
        rip = str(r.get("RiPort") or "")
        checks = [
            ("CellType", ct == "TDD", ct, "TDD"),
            ("RRU", rru.startswith("AAU"), rru, "AAU…"),
            ("RRUname", rruname == "A9651A_S26", rruname, "A9651A_S26"),
            ("rruPort", rrup == "1-64", rrup, "1-64"),
            ("RiPort", rip.startswith("VBP_1_3") or rip.startswith("VBP_1_8"), rip, "VBP_1_3… or VBP_1_8…"),
        ]
        for col, ok, actual, expected in checks:
            if not ok:
                findings.append(Finding(
                    sheet=r.sheet, row=r.row, column=col, ne_name=ne,
                    rule_id="r_5g_tdd_consistency", severity="error",
                    actual=actual, expected=expected,
                    message=f"Radio 5G expects {col} {expected!r}, got {actual!r}",
                ))
    return findings


def r_aau_bw_sum(ip: dict, rows4g: list[RadioRow], rows5g: list[RadioRow]) -> list[Finding]:
    """For each AAU (NE_Name + sector), the sum of Radio 4G TDD
    dlChannelBandwidth + Radio 5G bSChannelBwDL must equal 100."""
    g4: dict[tuple[str, int], list[tuple[int, float]]] = {}
    for r in rows4g:
        if r.get("arfcndl") != r.get("arfcnul") or r.get("arfcndl") is None:
            continue
        ne = normalize_to_ip_ne(str(r.get("NE_Name") or ""), ip, from_4g=True)
        sec = sector_from_cellname(str(r.get("CellName") or ""))
        bw = r.get("dlChannelBandwidth")
        if ne and sec is not None and isinstance(bw, (int, float)):
            g4.setdefault((ne, sec), []).append((r.row, float(bw)))

    g5: dict[tuple[str, int], list[tuple[int, float]]] = {}
    for r in rows5g:
        ne = normalize_to_ip_ne(str(r.get("NE_Name") or ""), ip, from_4g=False)
        sec = sector_from_rru(r.get("RRU"))
        bw = r.get("bSChannelBwDL")
        if ne and sec is not None and isinstance(bw, (int, float)):
            g5.setdefault((ne, sec), []).append((r.row, float(bw)))

    findings: list[Finding] = []
    for key in sorted(set(g4) | set(g5)):
        ne, sec = key
        sum_4g = sum(b for _, b in g4.get(key, []))
        sum_5g = sum(b for _, b in g5.get(key, []))
        total = sum_4g + sum_5g
        if total == 100:
            continue
        if g5.get(key):
            sheet, row, col = "Radio 5G", g5[key][0][0], "bSChannelBwDL"
        else:
            sheet, row, col = "Radio 4G", g4[key][0][0], "dlChannelBandwidth"
        findings.append(Finding(
            sheet=sheet, row=row, column=col, ne_name=ne,
            rule_id="r_aau_bw_sum", severity="error",
            actual=total, expected=100,
            message=(f"AAU {ne} sector {sec}: sum 4G dlChannelBandwidth ({sum_4g}) "
                     f"+ 5G bSChannelBwDL ({sum_5g}) = {total}, must be 100"),
        ))
    return findings


_R5G_INHERITED_TDD_TARGETS = {
    "bSChannelBwDL": 60,
    "configuredMaxTxPower": 120,
    "arfcnDL": 505998,
    "arfcnUL": 505998,
}


def r_5g_inherits_4g_tdd(rows4g: list[RadioRow], rows5g: list[RadioRow]) -> list[Finding]:
    """If a NE_Name has TDD rows in Radio 4G (arfcndl==arfcnul), every Radio 5G
    row for that NE_Name must use the standardized values defined in
    _R5G_INHERITED_TDD_TARGETS."""
    findings: list[Finding] = []
    tdd_g_ne: set[str] = set()
    for r in rows4g:
        if r.get("arfcndl") == r.get("arfcnul") and r.get("arfcndl") is not None:
            ip_ne = ip_ne_for_4g(str(r.get("NE_Name") or ""))
            if ip_ne:
                tdd_g_ne.add(ip_ne)

    for r in rows5g:
        ne = str(r.get("NE_Name") or "")
        if ne not in tdd_g_ne:
            continue
        for col, expected in _R5G_INHERITED_TDD_TARGETS.items():
            actual = r.get(col)
            if actual != expected:
                findings.append(Finding(
                    sheet=r.sheet, row=r.row, column=col, ne_name=ne,
                    rule_id="r_5g_inherits_4g_tdd", severity="error",
                    actual=actual, expected=expected,
                    message=f"NE_Name {ne} has TDD cells in 4G; 5G {col} must be {expected!r}, got {actual!r}",
                ))
    return findings


# ------------------------- output -------------------------

def render_console(findings: list[Finding]) -> str:
    if not findings:
        return "OK: no findings."
    lines = [f"Found {len(findings)} violation(s):", ""]
    by_sheet: dict[str, list[Finding]] = {}
    for f in findings:
        by_sheet.setdefault(f.sheet, []).append(f)
    for sheet, fs in by_sheet.items():
        lines.append(f"== {sheet} ({len(fs)}) ==")
        fs.sort(key=lambda f: (f.ne_name, f.row, f.column))
        for f in fs:
            lines.append(
                f"  row {f.row} [{f.ne_name}] {f.rule_id}: {f.column} = {f.actual!r}"
                f" -> expected {f.expected!r}"
            )
            lines.append(f"      {f.message}")
        lines.append("")
    return "\n".join(lines)


def write_json(findings: list[Finding], input_path: Path, out_path: Path) -> None:
    payload = {
        "input": str(input_path),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "count": len(findings),
        "findings": [asdict(f) for f in findings],
    }
    out_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def write_csv(findings: list[Finding], out_path: Path) -> None:
    fields = ["sheet", "row", "column", "ne_name", "rule_id", "severity", "actual", "expected", "message"]
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for finding in findings:
            row = asdict(finding)
            row["actual"] = "" if row["actual"] is None else str(row["actual"])
            row["expected"] = "" if row["expected"] is None else str(row["expected"])
            w.writerow(row)


# ------------------------- fix engine -------------------------

COL_INDEX = {"Radio 4G": R4G, "Radio 5G": R5G}


def collect_findings(ip, rows4g, rows5g) -> list[Finding]:
    findings: list[Finding] = []
    findings += r_ne_name_4g(ip, rows4g)
    findings += r_ne_name_5g(ip, rows5g)
    findings += r_enb_id(ip, rows4g)
    findings += r_gnb_id(ip, rows5g)
    findings += r_4g_tdd_fdd(rows4g)
    findings += r_4g_tdd_power(rows4g)
    findings += r_4g_rru_aau_from_cellname(rows4g)
    findings += r_4g_rruname_riport_rru_mapping(rows4g)
    findings += r_4g_rruport_tdd(rows4g)
    findings += r_5g_tdd_consistency(rows5g)
    findings += r_5g_inherits_4g_tdd(rows4g, rows5g)
    findings += r_aau_bw_sum(ip, rows4g, rows5g)
    return findings


def fixer_for(f: Finding):
    """Return the value to write for this finding, or None if not auto-fixable."""
    rid = f.rule_id
    if rid in ("r_ne_name_4g", "r_ne_name_5g", "r_enb_id", "r_gnb_id"):
        return f.expected if f.expected not in (None, "") else None

    if rid == "r_4g_tdd_fdd":
        if f.column == "CellType":
            return f.expected  # 'TDD' or 'FDD'
        if f.column == "RRUname" and f.expected == "A9651A_S26":
            return "A9651A_S26"
        if f.column == "RiPort" and isinstance(f.actual, str):
            if f.actual.startswith("VBP_1_6"):
                return "VBP_1_3" + f.actual[len("VBP_1_6"):]
            if f.actual.startswith("VBP_1_3"):
                return "VBP_1_6" + f.actual[len("VBP_1_3"):]
        if f.column == "RRU" and isinstance(f.actual, str):
            expected_str = str(f.expected or "")
            if expected_str.startswith("AAU"):
                target = "AAU"
            elif expected_str.startswith("RRU"):
                target = "RRU"
            else:
                return None
            rest = f.actual
            for p in ("AAU", "RRU"):
                if rest.startswith(p):
                    rest = rest[len(p):]
                    break
            rest = rest.lstrip("-_")
            return f"{target}_{rest}"
        return None

    if rid == "r_4g_rru_aau_from_cellname":
        return f.expected  # 'AAU_1' / 'AAU_2' / 'AAU_3' (deterministic)

    if rid == "r_4g_tdd_power":
        return f.expected

    if rid == "r_4g_rruname_band_mapping":
        return f.expected

    if rid == "r_4g_rruport_tdd":
        return "1-64"

    if rid == "r_5g_tdd_consistency":
        if f.column == "CellType" and f.expected == "TDD":
            return "TDD"
        if f.column == "RRUname" and f.expected == "A9651A_S26":
            return "A9651A_S26"
        if f.column == "rruPort" and f.expected == "1-64":
            return "1-64"
        return None

    if rid == "r_5g_inherits_4g_tdd":
        return f.expected  # specific constants from _R5G_INHERITED_TDD_TARGETS

    return None


def apply_fixes(wb, findings: list[Finding]) -> tuple[list[tuple[Finding, object]], list[Finding]]:
    applied: list[tuple[Finding, object]] = []
    skipped: list[Finding] = []
    for f in findings:
        new_value = fixer_for(f)
        cols = COL_INDEX.get(f.sheet)
        if new_value is None or cols is None or f.column not in cols:
            skipped.append(f)
            continue
        ws = wb[f.sheet]
        ws.cell(row=f.row, column=cols[f.column] + 1, value=new_value)
        applied.append((f, new_value))
    return applied, skipped


# ------------------------- main -------------------------

def _load(xlsx_path: Path, *, data_only: bool):
    wb = openpyxl.load_workbook(xlsx_path, data_only=data_only)
    if "IP" not in wb.sheetnames:
        print("ERROR: workbook has no 'IP' sheet", file=sys.stderr)
        return None
    return wb


def _gather(wb):
    ip = parse_ip(wb["IP"])
    rows4g = parse_radio(wb["Radio 4G"], "Radio 4G", R4G) if "Radio 4G" in wb.sheetnames else []
    rows5g = parse_radio(wb["Radio 5G"], "Radio 5G", R5G) if "Radio 5G" in wb.sheetnames else []
    return ip, rows4g, rows5g


def run(xlsx_path: Path) -> int:
    wb = _load(xlsx_path, data_only=True)
    if wb is None:
        return 2
    ip, rows4g, rows5g = _gather(wb)
    findings = collect_findings(ip, rows4g, rows5g)

    print(render_console(findings))

    json_path = xlsx_path.with_suffix(xlsx_path.suffix + ".findings.json")
    csv_path = xlsx_path.with_suffix(xlsx_path.suffix + ".findings.csv")
    write_json(findings, xlsx_path, json_path)
    write_csv(findings, csv_path)
    print(f"\nWrote {json_path}")
    print(f"Wrote {csv_path}")

    return 0 if not findings else 1


def run_fix(xlsx_path: Path) -> int:
    wb = _load(xlsx_path, data_only=False)
    if wb is None:
        return 2
    ip, rows4g, rows5g = _gather(wb)
    findings = collect_findings(ip, rows4g, rows5g)
    applied, skipped = apply_fixes(wb, findings)

    out_path = xlsx_path.with_name(xlsx_path.stem + "_fixed" + xlsx_path.suffix)
    wb.save(out_path)

    # Re-validate the corrected workbook in-memory
    ip2, rows4g2, rows5g2 = _gather(wb)
    remaining = collect_findings(ip2, rows4g2, rows5g2)

    print(f"Applied {len(applied)} fix(es); {len(skipped)} finding(s) not auto-fixable.")
    print(f"Saved fixed workbook: {out_path}")
    print()
    print("Residual findings (after fix):")
    print(render_console(remaining))

    json_path = out_path.with_suffix(out_path.suffix + ".findings.json")
    csv_path = out_path.with_suffix(out_path.suffix + ".findings.csv")
    write_json(remaining, out_path, json_path)
    write_csv(remaining, csv_path)
    print(f"\nWrote {json_path}")
    print(f"Wrote {csv_path}")

    return 0 if not remaining else 1


def main(argv: list[str]) -> int:
    args = argv[1:]
    fix_mode = False
    if args and args[0] == "--fix":
        fix_mode = True
        args = args[1:]
    if len(args) != 1:
        print("Usage: verify_cdd.py [--fix] <xlsx-path>", file=sys.stderr)
        return 2
    path = Path(args[0])
    return run_fix(path) if fix_mode else run(path)


if __name__ == "__main__":
    sys.exit(main(sys.argv))
