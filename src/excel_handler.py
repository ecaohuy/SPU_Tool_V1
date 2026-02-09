"""Excel file handling for SPU Processing Tool."""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os


class ExcelHandler:
    """Handle Excel file operations."""

    # Standard sheets to read from CDD input file
    STANDARD_SHEETS = [
        "IP", "Radio 2G", "Radio 3G", "Radio 4G", "Radio 5G",
        "2G-2G", "2G-3G", "2G-4G", "3G-2G", "3G-3G", "3G-4G",
        "RET", "Mapping"
    ]

    def __init__(self):
        self.input_data = {}
        self.template_workbook = None
        self.template_sheets = []

    def read_input_file(self, file_path):
        """Read CDD input Excel file and return data for all sheets.

        Args:
            file_path: Path to the CDD input Excel file

        Returns:
            dict: Dictionary with sheet names as keys and DataFrames as values
        """
        self.input_data = {}

        try:
            # Get all sheet names from the file
            xl = pd.ExcelFile(file_path)
            available_sheets = xl.sheet_names

            # Read each standard sheet if it exists
            for sheet_name in self.STANDARD_SHEETS:
                if sheet_name in available_sheets:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        # Filter out instruction/header rows
                        df = self._filter_instruction_rows(df, sheet_name)
                        self.input_data[sheet_name] = df
                    except Exception as e:
                        print(f"Warning: Could not read sheet '{sheet_name}': {e}")
                        self.input_data[sheet_name] = pd.DataFrame()
                else:
                    self.input_data[sheet_name] = pd.DataFrame()

            return self.input_data

        except Exception as e:
            raise Exception(f"Failed to read input file: {e}")

    def _filter_instruction_rows(self, df, sheet_name):
        """Filter out instruction/header rows from the dataframe.

        The CDD input files have 3 instruction rows at the top:
        - Row 1: Column descriptions (e.g., "Site name")
        - Row 2: Data types (e.g., "string:[1..255]")
        - Row 3: Required indicators (e.g., "Mandatory")

        Args:
            df: DataFrame to filter
            sheet_name: Name of the sheet

        Returns:
            pd.DataFrame: Filtered DataFrame with only data rows
        """
        if df.empty:
            return df

        # For sheets with NE_Name column, filter by valid NE_Name pattern
        if "NE_Name" in df.columns:
            # Valid NE_Name starts with 'g' or 'e' followed by letters and numbers
            # e.g., gBL00231Z, eCM00025Z, gBLT8509, gCMT8910
            mask = df["NE_Name"].astype(str).str.match(r'^[ge][A-Z]{2,3}\d+[A-Z]?$', na=False)
            return df[mask].reset_index(drop=True)

        # For other sheets, try to filter by common patterns
        # Skip rows where first column contains instruction text
        first_col = df.columns[0]
        if first_col in df.columns:
            # Filter out rows with common instruction patterns
            instruction_patterns = [
                'Mandatory', 'Optional', 'string:', 'integer:',
                'Bắt buộc', 'Có thể', 'use for', 'Site name'
            ]
            mask = ~df[first_col].astype(str).str.contains(
                '|'.join(instruction_patterns),
                case=False,
                na=False
            )
            return df[mask].reset_index(drop=True)

        return df

    def read_template_file(self, file_path):
        """Read SPU template Excel file.

        Args:
            file_path: Path to the SPU template Excel file

        Returns:
            openpyxl.Workbook: The template workbook
        """
        try:
            self.template_workbook = load_workbook(file_path)
            self.template_sheets = self.template_workbook.sheetnames
            return self.template_workbook
        except Exception as e:
            raise Exception(f"Failed to read template file: {e}")

    def write_output_file(self, output_path, data_dict):
        """Write processed data to output Excel file.

        Args:
            output_path: Path for the output file
            data_dict: Dictionary with sheet names as keys and DataFrames as values
        """
        try:
            # If template exists, copy it and modify
            if self.template_workbook:
                wb = self.template_workbook
            else:
                from openpyxl import Workbook
                wb = Workbook()

            for sheet_name, df in data_dict.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Clear existing data (keep headers)
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            cell.value = None
                else:
                    ws = wb.create_sheet(title=sheet_name)

                # Write DataFrame to sheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(output_path)

        except Exception as e:
            raise Exception(f"Failed to write output file: {e}")

    def get_sheet_data(self, sheet_name):
        """Get data for a specific sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            pd.DataFrame: DataFrame for the sheet, or empty DataFrame if not found
        """
        return self.input_data.get(sheet_name, pd.DataFrame())

    def get_all_sheet_names(self):
        """Get all available sheet names from input data.

        Returns:
            list: List of sheet names
        """
        return list(self.input_data.keys())

    def get_sheet_columns(self, sheet_name):
        """Get column names for a specific sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            list: List of column names
        """
        df = self.get_sheet_data(sheet_name)
        if not df.empty:
            return list(df.columns)
        return []

    def get_sheet_row_count(self, sheet_name):
        """Get row count for a specific sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            int: Number of rows (excluding header)
        """
        df = self.get_sheet_data(sheet_name)
        return len(df)
