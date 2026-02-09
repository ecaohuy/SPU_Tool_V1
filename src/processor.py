"""SPU Processing Logic for the Tool."""

import json
import math
import os
import re

import pandas as pd
from openpyxl import load_workbook

from .utils import get_config_path, ensure_output_folder, generate_output_filename


def safe_int(value, default=None):
    """Safely convert value to int, returning default if conversion fails."""
    if pd.isna(value):
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def safe_str(value, default=""):
    """Safely convert value to string, returning default if None/NaN."""
    if pd.isna(value):
        return default
    return str(value).strip()


class SPUProcessor:
    """Process CDD input data using SPU template and config mappings."""

    def __init__(self):
        self.config = None
        self.input_data = None
        self.template_workbook = None
        self.template_path = None
        self.mappings = {}  # Parsed mappings from Mapping sheet
        self.version = "V1.70.26"

    def load_config(self):
        """Load configuration from config.json."""
        config_path = get_config_path()
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
            return True
        except Exception as e:
            raise Exception(f"Failed to load config: {e}")

    def set_input_data(self, input_data):
        """Set the input data from Excel handler."""
        self.input_data = input_data
        if self.config is None:
            self.load_config()
        # Parse mappings from Mapping sheet
        if "Mapping" in input_data and not input_data["Mapping"].empty:
            self._parse_mappings(input_data["Mapping"])

    def _parse_mappings(self, mapping_df):
        """Parse the Mapping sheet into usable format."""
        self.mappings = {}

        # Filter by version
        df = mapping_df[mapping_df["Version"] == self.version]

        for _, row in df.iterrows():
            sheet = safe_str(row.get("Sheet"))
            column = safe_str(row.get("Column"))
            source_sheet = safe_str(row.get("SourceSheet"))
            source_column = safe_str(row.get("SourceColumn"))
            fixed_value = row.get("FixedValue")

            if not sheet or not column:
                continue

            if sheet not in self.mappings:
                self.mappings[sheet] = []

            self.mappings[sheet].append({
                "column": column,
                "source_sheet": source_sheet,
                "source_column": source_column,
                "fixed_value": fixed_value if pd.notna(fixed_value) else None
            })

    def set_template(self, template_path):
        """Set the template file path."""
        self.template_path = template_path
        try:
            self.template_workbook = load_workbook(template_path)
        except Exception as e:
            raise Exception(f"Failed to load template: {e}")

    def get_groups(self):
        """Get unique groups - combine all stations into one."""
        if "IP" in self.input_data and not self.input_data["IP"].empty:
            ip_df = self.input_data["IP"]
            if "NE_Name" in ip_df.columns:
                ne_names = ip_df["NE_Name"].dropna().tolist()
                if len(ne_names) >= 2:
                    return [f"{ne_names[0]}-{ne_names[-1]}"]
                elif len(ne_names) == 1:
                    return [ne_names[0]]
        return ["default"]

    def process(self, progress_callback=None):
        """Process the input data and generate output files."""
        if not self.input_data:
            raise Exception("No input data loaded")
        if not self.template_path:
            raise Exception("No template selected")
        if not self.config:
            self.load_config()

        output_files = []
        groups = self.get_groups()

        if progress_callback:
            progress_callback("Starting SPU processing...", 0)

        for idx, group in enumerate(groups):
            if progress_callback:
                pct = int((idx / len(groups)) * 100)
                progress_callback(f"Processing group: {group}", pct)

            output_path = self._process_group(group)
            output_files.append(output_path)

        if progress_callback:
            progress_callback("Processing complete!", 100)

        return output_files

    def _get_template_columns(self, ws):
        """Get column name to position mapping from template row 2."""
        columns = {}
        for col in range(1, 100):
            val = ws.cell(row=2, column=col).value
            if val:
                columns[val] = col
        return columns

    def _find_data_start_row(self, ws):
        """Find the first data row (after headers)."""
        for row in range(1, 10):
            cell_value = str(ws.cell(row=row, column=1).value or "")
            if "Primary Key" in cell_value:
                return row + 1
        return 6  # Default: assume 5 header rows

    def _process_group(self, group):
        """Process a single group and generate output file."""
        wb = load_workbook(self.template_path)

        # Process 'site' sheet first (special handling for subNetwork)
        self._process_site_sheet(wb)

        # Process other sheets using mappings
        self._process_sheet_with_mapping(wb, "RU", "Radio 4G")
        # Use specialized cell4g processing with calculated fields
        self._process_cell4g_sheet(wb)
        # Use specialized cell5g processing with calculated fields
        self._process_cell5g_sheet(wb)

        # Process cable sheet with special RiPort parsing
        self._process_cable_sheet(wb)

        # Special processing for sheets not in Mapping
        self._process_ip_sheet(wb)
        self._process_sctp_sheet(wb)
        self._process_bbu_sheet(wb)
        self._process_drycontact_sheet(wb)
        self._process_aisgctrlport_sheet(wb)

        # Generate output filename and save
        output_folder = ensure_output_folder()
        template_name = os.path.basename(self.template_path)
        output_filename = generate_output_filename(template_name, group)
        output_path = os.path.join(output_folder, output_filename)

        wb.save(output_path)
        return output_path

    def _process_site_sheet(self, wb):
        """Process 'site' sheet with special subNetwork mapping.

        Mapping rules:
        1. subNetwork = first 3 characters of NE_Name mapped via config.json province
        2. meId = NE_Name from IP sheet
        3. userLabel = NE_Name from IP sheet
        4. ipAddress = OAM_IP from IP sheet
        5. Other fields from Mapping sheet (fixed values)
        """
        if "site" not in wb.sheetnames:
            return

        ws = wb["site"]
        template_cols = self._get_template_columns(ws)
        ip_df = self.input_data.get("IP", pd.DataFrame())

        if ip_df.empty:
            return

        # Get province mapping from config
        province_mapping = self.config.get("province", {})

        # Get fixed values from Mapping sheet for 'site'
        site_fixed_values = {}
        if "site" in self.mappings:
            for mapping in self.mappings["site"]:
                if mapping["fixed_value"] is not None:
                    site_fixed_values[mapping["column"]] = mapping["fixed_value"]

        start_row = self._find_data_start_row(ws)

        # Process each row of IP data
        row = start_row
        for _, ip_row in ip_df.iterrows():
            ne_name = safe_str(ip_row.get("NE_Name", ""))
            if not ne_name:
                continue

            # 1. subNetwork = first 3 chars of NE_Name mapped to province
            prefix = ne_name[:3] if len(ne_name) >= 3 else ne_name
            sub_network = province_mapping.get(prefix, prefix)
            if "subNetwork" in template_cols:
                ws.cell(row=row, column=template_cols["subNetwork"], value=sub_network)

            # 2. meId = NE_Name
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=ne_name)

            # 3. userLabel = NE_Name
            if "userLabel" in template_cols:
                ws.cell(row=row, column=template_cols["userLabel"], value=ne_name)

            # 4. ipAddress = OAM_IP
            oam_ip = safe_str(ip_row.get("OAM_IP", ""))
            if "ipAddress" in template_cols:
                ws.cell(row=row, column=template_cols["ipAddress"], value=oam_ip)

            # 5. Apply fixed values from Mapping sheet
            for col_name, fixed_val in site_fixed_values.items():
                if col_name in template_cols:
                    ws.cell(row=row, column=template_cols[col_name], value=fixed_val)

            row += 1

    def _process_sheet_with_mapping(self, wb, target_sheet, primary_source):
        """Process a sheet using the Mapping sheet definitions."""
        if target_sheet not in wb.sheetnames:
            return
        if target_sheet not in self.mappings:
            return

        ws = wb[target_sheet]
        template_cols = self._get_template_columns(ws)
        mappings = self.mappings[target_sheet]

        # Get source data
        source_df = self.input_data.get(primary_source, pd.DataFrame())
        if source_df.empty:
            return

        start_row = self._find_data_start_row(ws)

        # For RU sheet, we need unique RRU entries
        if target_sheet == "RU":
            self._process_ru_with_mapping(ws, template_cols, mappings, start_row)
            return

        # Process each row of source data
        row = start_row
        for _, source_row in source_df.iterrows():
            for mapping in mappings:
                col_name = mapping["column"]
                if col_name not in template_cols:
                    continue

                col_pos = template_cols[col_name]
                value = self._get_mapped_value(mapping, source_row)

                if value is not None:
                    ws.cell(row=row, column=col_pos, value=value)
            row += 1

    def _process_ru_with_mapping(self, ws, template_cols, mappings, start_row):
        """Process RU sheet with unique RRU entries.

        Mapping rules:
        - meId = NE_Name from IP sheet
        - moId = RRU column from Radio 4G/5G (deduplicated)
        - name = RRUname column from Radio 4G/5G
        - hwWorkScence = mapped from hwWorkScence_mapping based on RRUname(&RRU) and technology
        - functionMode = mapped from functionMode_mapping based on RRUname(&RRU) and technology
        - sectorFunction = CellNames joined by '&' for all cells using that RRU
        - RxChannelNo = mapping RRU and rruPort (e.g., RRU-1 with 2 rruPorts 1234 -> 1-4&1-4)
        - TxChannelNo = same as RxChannelNo
        - sectorFreqPower = band (from arfcnDL via bandIndicator_mapping) & maxCPTransPwr
        - sharedSwitch = fixed value from Mapping sheet
        - networkingType = fixed value from Mapping sheet
        """
        radio4g = self.input_data.get("Radio 4G", pd.DataFrame())
        radio5g = self.input_data.get("Radio 5G", pd.DataFrame())

        spu_config = self.config.get("SPU", {}).get(self.version, {})
        hw_mapping = spu_config.get("hwWorkScence_mapping", {})
        func_mapping = spu_config.get("functionMode_mapping", {})
        band_indicator_mapping = spu_config.get("bandIndicator_mapping", {})
        rru_port_mapping = spu_config.get("rruPort_mapping", {})

        # Get fixed values from Mapping sheet for RU
        ru_fixed_values = {}
        if "RU" in self.mappings:
            for mapping in self.mappings["RU"]:
                if mapping["fixed_value"] is not None:
                    ru_fixed_values[mapping["column"]] = mapping["fixed_value"]

        # Build a dictionary of all RRUs with their technologies, cell names, rruPorts, and sectorFreqPower info
        # Key: (ne_name, rru_name), Value: dict with rru_type, techs, cells, rru_ports, sector_freq_power_info
        rru_info = {}

        # Process 4G RRUs
        if not radio4g.empty and "RRU" in radio4g.columns:
            for _, source_row in radio4g.iterrows():
                ne_name = safe_str(source_row.get("NE_Name", ""))
                rru_name = safe_str(source_row.get("RRU", ""))
                rru_type = safe_str(source_row.get("RRUname", ""))
                cell_name = safe_str(source_row.get("CellName", ""))
                rru_port = safe_str(source_row.get("rruPort", ""))
                arfcn_dl = source_row.get("arfcndl", "")
                max_cp_trans_pwr = source_row.get("maxCPTransPwr", "")

                if not ne_name or not rru_name:
                    continue

                key = (ne_name, rru_name)
                if key not in rru_info:
                    rru_info[key] = {
                        'rru_type': rru_type,
                        'techs': set(),
                        'cells_4g': [],
                        'cells_5g': [],
                        'rru_ports': [],  # Store ALL rruPort occurrences (not unique)
                        'sector_freq_power_info': []  # Store ALL sector freq power entries
                    }
                rru_info[key]['techs'].add('4G')
                if cell_name and cell_name not in rru_info[key]['cells_4g']:
                    rru_info[key]['cells_4g'].append(cell_name)
                # Collect ALL rruPort occurrences for RxChannelNo/TxChannelNo (not unique)
                # e.g., RRU-1 with 2 cells using rruPort 1234 -> [1234, 1234] -> 1-4&1-4
                if rru_port:
                    rru_info[key]['rru_ports'].append((rru_port, '4G'))
                # Collect arfcnDL and maxCPTransPwr for sectorFreqPower (ALL entries, not unique)
                if arfcn_dl:
                    arfcn_dl_str = str(safe_int(arfcn_dl, 0))
                    band = band_indicator_mapping.get(arfcn_dl_str, "")
                    pwr = safe_int(max_cp_trans_pwr, "")
                    if band and pwr:
                        freq_power_entry = f"{band}:{pwr}"
                        rru_info[key]['sector_freq_power_info'].append(freq_power_entry)

        # Process 5G RRUs/AAUs
        if not radio5g.empty and "RRU" in radio5g.columns:
            for _, source_row in radio5g.iterrows():
                ne_name = safe_str(source_row.get("NE_Name", ""))
                rru_name = safe_str(source_row.get("RRU", ""))
                rru_type = safe_str(source_row.get("RRUname", ""))
                # For 5G, CellName is in 'nRCell' column
                cell_name = safe_str(source_row.get("nRCell", ""))
                rru_port = safe_str(source_row.get("rruPort", ""))
                # For 5G sectorFreqPower: bandListManual & configuredMaxTxPower
                band_list_manual = source_row.get("bandListManual", "")
                configured_max_tx_pwr = source_row.get("configuredMaxTxPower", "")

                if not ne_name or not rru_name:
                    continue

                key = (ne_name, rru_name)
                if key not in rru_info:
                    rru_info[key] = {
                        'rru_type': rru_type,
                        'techs': set(),
                        'cells_4g': [],
                        'cells_5g': [],
                        'rru_ports': [],
                        'sector_freq_power_info': []
                    }
                rru_info[key]['techs'].add('5G')
                if cell_name and cell_name not in rru_info[key]['cells_5g']:
                    rru_info[key]['cells_5g'].append(cell_name)
                # Collect ALL rruPort occurrences for RxChannelNo/TxChannelNo (not unique)
                # e.g., AAU_1 with 3 cells using rruPort 1-64 -> 1-64&1-64&1-64
                if rru_port:
                    rru_info[key]['rru_ports'].append((rru_port, '5G'))
                # Collect bandListManual and configuredMaxTxPower for sectorFreqPower (5G)
                if band_list_manual and pd.notna(configured_max_tx_pwr):
                    pwr_5g = safe_int(configured_max_tx_pwr, "")
                    if pwr_5g:
                        freq_power_entry = f"{band_list_manual}:{pwr_5g}"
                        rru_info[key]['sector_freq_power_info'].append(freq_power_entry)
                # Update rru_type if not set (in case RRU appears in 5G but not 4G)
                if not rru_info[key]['rru_type']:
                    rru_info[key]['rru_type'] = rru_type

        # Write RU entries to worksheet
        row = start_row
        for (ne_name, rru_name), info in rru_info.items():
            rru_type = info['rru_type']
            techs = info['techs']
            cells_4g = info['cells_4g']
            cells_5g = info['cells_5g']
            rru_ports = info['rru_ports']
            sector_freq_power_info = info['sector_freq_power_info']

            # Determine technology key for mapping
            if '4G' in techs and '5G' in techs:
                tech_key = "4G,5G"
            elif '5G' in techs:
                tech_key = "5G"
            else:
                tech_key = "4G"

            # Build sectorFunction from cell names
            # Combine 4G and 5G cells, joined by '&'
            all_cells = cells_4g + cells_5g
            sector_function = "&".join(all_cells) if all_cells else ""

            # Build RxChannelNo/TxChannelNo from rruPorts
            # Map rruPort values: 1234 -> 1-4, 12 -> 1-2, 34 -> 3-4, 1-64 -> 1-64, etc.
            rx_channel_parts = []
            for rru_port, _ in rru_ports:
                # Use rruPort_mapping from config if available
                mapped_port = rru_port_mapping.get(str(rru_port), "")
                if not mapped_port:
                    # Apply default mapping logic
                    if rru_port == "1234":
                        mapped_port = "1-4"
                    elif rru_port == "12":
                        mapped_port = "1-2"
                    elif rru_port == "34":
                        mapped_port = "3-4"
                    elif rru_port in ["1-64", "1-32"]:
                        mapped_port = rru_port
                    elif rru_port == "1":
                        mapped_port = "1"
                    elif rru_port == "2":
                        mapped_port = "2"
                    elif rru_port == "3":
                        mapped_port = "3"
                    elif rru_port == "4":
                        mapped_port = "4"
                    else:
                        mapped_port = rru_port
                rx_channel_parts.append(mapped_port)
            rx_channel_no = "&".join(rx_channel_parts) if rx_channel_parts else ""
            tx_channel_no = rx_channel_no  # TxChannelNo = RxChannelNo

            # Build sectorFreqPower: band:maxCPTransPwr joined by '&'
            sector_freq_power = "&".join(sector_freq_power_info) if sector_freq_power_info else ""

            # Write to worksheet
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=ne_name)
            if "moId" in template_cols:
                ws.cell(row=row, column=template_cols["moId"], value=rru_name)
            if "name" in template_cols:
                ws.cell(row=row, column=template_cols["name"], value=rru_type)
            if "hwWorkScence" in template_cols:
                # Use RRUname (without & prefix) for mapping lookup
                # Extract the device type from rru_type (e.g., R9264S_M1821(AAA))
                hw = hw_mapping.get(rru_type, {}).get(tech_key, "")
                ws.cell(row=row, column=template_cols["hwWorkScence"], value=hw)
            if "functionMode" in template_cols:
                fm = func_mapping.get(rru_type, {}).get(tech_key, "")
                ws.cell(row=row, column=template_cols["functionMode"], value=fm)
            if "sectorFunction" in template_cols:
                ws.cell(row=row, column=template_cols["sectorFunction"], value=sector_function)
            if "RxChannelNo" in template_cols:
                ws.cell(row=row, column=template_cols["RxChannelNo"], value=rx_channel_no)
            if "TxChannelNo" in template_cols:
                ws.cell(row=row, column=template_cols["TxChannelNo"], value=tx_channel_no)
            if "sectorFreqPower" in template_cols:
                ws.cell(row=row, column=template_cols["sectorFreqPower"], value=sector_freq_power)

            # Apply fixed values from Mapping sheet (sharedSwitch, networkingType, etc.)
            already_set = ["meId", "moId", "name", "hwWorkScence", "functionMode",
                          "sectorFunction", "RxChannelNo", "TxChannelNo", "sectorFreqPower"]
            for col_name, fixed_val in ru_fixed_values.items():
                if col_name in template_cols and col_name not in already_set:
                    ws.cell(row=row, column=template_cols[col_name], value=fixed_val)

            row += 1

    def _process_cell5g_sheet(self, wb):
        """Process cell5g sheet with specialized calculated fields.

        Mapping rules (from user specification):
        - meId: NE_Name from Radio 5G
        - radioTemplate5G: fixed value from Mapping sheet
        - functionMoId: 452-04_gNBId (e.g., 452-04_12940231)
        - functionUserLabel: NE_Name
        - gNBId: gNBId from Radio 5G
        - gNBIdLength: fixed 26
        - cellLocalId: cellLocalId from Radio 5G
        - userLabel: nRCell from Radio 5G
        - NSSAI: fixed value from Mapping sheet
        - masterOperatorId: 45204-gNBId-cellLocalId (e.g., 45204-12940231-111)
        - pci: nRPCI
        - tac: nRTAC
        - cellAtt: fixed value (sub6G)
        - duplexMode: CellType from Radio 5G
        - coverageType: fixed value (Macro)
        - qcellFlag: fixed 0
        - superMIMOSwitch: fixed 0
        - NRCarrierId: NRCarrierId from Radio 5G
        - frequencyBandListUL/DL: bandListManual from Radio 5G
        - frequencyDL: 0.005 * arfcnDL
        - frequencyUL: 0.005 * arfcnUL
        - gSCNOffset: fixed 1
        - nrCellScene: fixed Normal
        - dlAntNum/ulAntNum: 64 if rruPort is 1-64, 5 if rruPort is 1-32
        - bpPoolFunctionId: extracted from RiPort Baseband (VBP_1_3&OF1 -> BF_3)
        - sectorFunctionId: nRCell
        - configuredMaxTxPower: 10 * log10(configuredMaxTxPower_watts)
        - nrbandwidth: bSChannelBwDL
        - prachRootSequenceValue: rachRootSequence
        - NRFreqRelation_ssbFrequency: 0.005 * ssbFrequency
        - NRFreqRelation_freqBandIndicator: bandListManual
        - Other fields: fixed values from Mapping sheet
        """
        if "cell5g" not in wb.sheetnames:
            return

        ws = wb["cell5g"]
        template_cols = self._get_template_columns(ws)

        radio5g = self.input_data.get("Radio 5G", pd.DataFrame())
        if radio5g.empty:
            return

        # Get fixed values from Mapping sheet for cell5g
        cell5g_fixed_values = {}
        if "cell5g" in self.mappings:
            for mapping in self.mappings["cell5g"]:
                if mapping["fixed_value"] is not None:
                    cell5g_fixed_values[mapping["column"]] = mapping["fixed_value"]

        start_row = self._find_data_start_row(ws)
        row = start_row

        for _, source_row in radio5g.iterrows():
            # Extract source values
            ne_name = safe_str(source_row.get("NE_Name", ""))
            gnb_id = safe_str(source_row.get("gNBId", ""))
            cell_local_id = safe_str(source_row.get("cellLocalId", ""))
            nrcell = safe_str(source_row.get("nRCell", ""))
            nr_pci = source_row.get("nRPCI", "")
            nr_tac = source_row.get("nRTAC", "")
            cell_type = safe_str(source_row.get("CellType", ""))
            nr_carrier_id = source_row.get("NRCarrierId", "")
            band_list_manual = source_row.get("bandListManual", "")
            arfcn_dl = source_row.get("arfcnDL", 0)
            arfcn_ul = source_row.get("arfcnUL", 0)
            ssb_frequency = source_row.get("ssbFrequency", 0)
            rru_port = safe_str(source_row.get("rruPort", ""))
            ri_port_baseband = safe_str(source_row.get("RiPort Baseband", ""))
            configured_max_tx_power = source_row.get("configuredMaxTxPower", 0)
            bs_channel_bw_dl = source_row.get("bSChannelBwDL", "")
            rach_root_sequence = source_row.get("rachRootSequence", "")

            # Direct mappings from Radio 5G
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=ne_name)
            if "functionUserLabel" in template_cols:
                ws.cell(row=row, column=template_cols["functionUserLabel"], value=ne_name)
            if "gNBId" in template_cols:
                ws.cell(row=row, column=template_cols["gNBId"], value=gnb_id)
            if "cellLocalId" in template_cols:
                ws.cell(row=row, column=template_cols["cellLocalId"], value=cell_local_id)
            if "userLabel" in template_cols:
                ws.cell(row=row, column=template_cols["userLabel"], value=nrcell)
            if "nrPhysicalCellDUId" in template_cols:
                ws.cell(row=row, column=template_cols["nrPhysicalCellDUId"], value=nr_carrier_id)
            if "nrPhysicalCellUserLabel" in template_cols:
                ws.cell(row=row, column=template_cols["nrPhysicalCellUserLabel"], value=nrcell)
            if "pci" in template_cols:
                ws.cell(row=row, column=template_cols["pci"], value=nr_pci)
            if "tac" in template_cols:
                ws.cell(row=row, column=template_cols["tac"], value=nr_tac)
            if "duplexMode" in template_cols:
                ws.cell(row=row, column=template_cols["duplexMode"], value=cell_type)
            if "NRCarrierId" in template_cols:
                ws.cell(row=row, column=template_cols["NRCarrierId"], value=nr_carrier_id)
            if "frequencyBandListUL" in template_cols:
                ws.cell(row=row, column=template_cols["frequencyBandListUL"], value=band_list_manual)
            if "frequencyBandListDL" in template_cols:
                ws.cell(row=row, column=template_cols["frequencyBandListDL"], value=band_list_manual)
            if "sectorFunctionId" in template_cols:
                ws.cell(row=row, column=template_cols["sectorFunctionId"], value=nrcell)
            if "nrbandwidth" in template_cols:
                ws.cell(row=row, column=template_cols["nrbandwidth"], value=bs_channel_bw_dl)
            if "prachRootSequenceValue" in template_cols:
                ws.cell(row=row, column=template_cols["prachRootSequenceValue"], value=rach_root_sequence)
            if "NRFreqRelation_freqBandIndicator" in template_cols:
                ws.cell(row=row, column=template_cols["NRFreqRelation_freqBandIndicator"], value=band_list_manual)

            # Calculated fields
            # functionMoId = 452-04_gNBId
            if "functionMoId" in template_cols:
                function_mo_id = f"452-04_{gnb_id}"
                ws.cell(row=row, column=template_cols["functionMoId"], value=function_mo_id)

            # masterOperatorId = 45204-gNBId-cellLocalId
            if "masterOperatorId" in template_cols:
                master_operator_id = f"45204-{gnb_id}-{cell_local_id}"
                ws.cell(row=row, column=template_cols["masterOperatorId"], value=master_operator_id)

            # frequencyDL = 0.005 * arfcnDL
            if "frequencyDL" in template_cols:
                try:
                    freq_dl = 0.005 * float(arfcn_dl) if pd.notna(arfcn_dl) else 0
                    ws.cell(row=row, column=template_cols["frequencyDL"], value=freq_dl)
                except (ValueError, TypeError):
                    pass

            # frequencyUL = 0.005 * arfcnUL
            if "frequencyUL" in template_cols:
                try:
                    freq_ul = 0.005 * float(arfcn_ul) if pd.notna(arfcn_ul) else 0
                    ws.cell(row=row, column=template_cols["frequencyUL"], value=freq_ul)
                except (ValueError, TypeError):
                    pass

            # NRFreqRelation_ssbFrequency = 0.005 * ssbFrequency
            if "NRFreqRelation_ssbFrequency" in template_cols:
                try:
                    ssb_freq = 0.005 * float(ssb_frequency) if pd.notna(ssb_frequency) else 0
                    ws.cell(row=row, column=template_cols["NRFreqRelation_ssbFrequency"], value=ssb_freq)
                except (ValueError, TypeError):
                    pass

            # dlAntNum/ulAntNum: 6 if rruPort is 1-64, 5 if rruPort is 1-32
            ant_num = 6  # default
            if rru_port == "1-64":
                ant_num = 6
            elif rru_port == "1-32":
                ant_num = 5
            if "dlAntNum" in template_cols:
                ws.cell(row=row, column=template_cols["dlAntNum"], value=ant_num)
            if "ulAntNum" in template_cols:
                ws.cell(row=row, column=template_cols["ulAntNum"], value=ant_num)

            # bpPoolFunctionId: extract from RiPort Baseband (VBP_1_3&OF1 -> BF_3)
            if "bpPoolFunctionId" in template_cols:
                bp_pool_func_id = ""
                if ri_port_baseband:
                    match = re.search(r'VBP_1_(\d+)&', ri_port_baseband)
                    if match:
                        bp_pool_func_id = f"BF_{match.group(1)}"
                ws.cell(row=row, column=template_cols["bpPoolFunctionId"], value=bp_pool_func_id)

            # configuredMaxTxPower = 10 * (10 * log10(power_input) + 30)
            if "configuredMaxTxPower" in template_cols:
                try:
                    power_input = float(configured_max_tx_power) if pd.notna(configured_max_tx_power) else 0
                    if power_input > 0:
                        power_output = int(10 * (10 * math.log10(power_input) + 30))
                    else:
                        power_output = 0
                    ws.cell(row=row, column=template_cols["configuredMaxTxPower"], value=power_output)
                except (ValueError, TypeError):
                    pass

            # powerPerRERef is fixed 178 for all meId
            if "powerPerRERef" in template_cols:
                ws.cell(row=row, column=template_cols["powerPerRERef"], value=178)

            # Apply fixed values from Mapping sheet
            for col_name, fixed_val in cell5g_fixed_values.items():
                if col_name in template_cols:
                    # Skip fields that have been calculated or directly mapped
                    already_set = [
                        "meId", "functionUserLabel", "gNBId", "cellLocalId", "userLabel",
                        "nrPhysicalCellDUId", "nrPhysicalCellUserLabel", "pci", "tac",
                        "duplexMode", "NRCarrierId", "frequencyBandListUL", "frequencyBandListDL",
                        "sectorFunctionId", "nrbandwidth", "prachRootSequenceValue",
                        "NRFreqRelation_freqBandIndicator", "functionMoId", "masterOperatorId",
                        "frequencyDL", "frequencyUL", "NRFreqRelation_ssbFrequency",
                        "dlAntNum", "ulAntNum", "bpPoolFunctionId", "configuredMaxTxPower",
                        "powerPerRERef"
                    ]
                    if col_name not in already_set:
                        ws.cell(row=row, column=template_cols[col_name], value=fixed_val)

            row += 1

    def _get_mapped_value(self, mapping, source_row):
        """Get value for a mapping from source row or fixed value."""
        # Fixed value takes priority
        if mapping["fixed_value"] is not None:
            return mapping["fixed_value"]

        # Get from source column
        source_col = mapping["source_column"]
        if source_col and source_col in source_row.index:
            value = source_row[source_col]
            if pd.notna(value):
                return value

        return None

    def _process_cell4g_sheet(self, wb):
        """Process cell4g sheet with specialized calculated fields.

        Mapping rules (from user specification):
        - meId: NE_Name from Radio 4G (IP sheet)
        - radioTemplate4G: fixed value from Mapping sheet (radio_lte_cudu_cell_template)
        - functionMoId: plmn_eNBId (e.g., 452-04_60486)
        - functionUserLabel: meId (NE_Name)
        - eNBId: eNBId from Radio 4G (IP sheet)
        - eNBIdLength: fixed value from Mapping sheet
        - cellLocalId: cellLocalId (cellId) from Radio 4G
        - userLabel: CellName from Radio 4G
        - radioMode: CellType from Radio 4G (FDD/TDD)
        - pci: PCI from Radio 4G
        - tac: TAC from Radio 4G
        - bandIndicator: mapped from arfcndl via bandIndicator_mapping in config
        - earfcnDl: arfcndl from Radio 4G
        - earfcnUl: mapped from arfcndl via earfcn_mapping in config
        - bandWidthDl: mapped from dlChannelBandwidth (20->5, 15->4, 10->3)
        - bandWidthUl: mapped from ulChannelBandwidth (same logic)
        - bandWidth (TDD): mapped from dlChannelBandwidth for TDD cells
        - rootSequenceIndex: RSI from Radio 4G
        - cpSpeRefSigPwr: cpSpeRefSigPwr from Radio 4G
        - paForDTCH: 4 if FDD, 2 if TDD
        - cellRSPortNum: mapped from MIMO (2T2R->1, 4T4R->2)
        - sampleRateCfg: 2 if bandWidthDl >= 4, else 0
        - bpPoolFunctionId: extracted from RiPort Baseband (VBP_1_6&OF1 -> BF_6)
        - sectorFunctionId: CellName
        - anttoPortMap: mapped based on rruPort and CellType
        - refPlmn, sceneCfg, urgencyeai, cellSize, mmModeSwch, pb, ranSharSwch,
          dedctRANSharingSwch, encrypAlgPriority, rfAppMode: fixed values from config
        - nRearfcn: NRFreqRelation_ssbFrequency from cell5g (for 4G-5G relation)
        """
        if "cell4g" not in wb.sheetnames:
            return

        ws = wb["cell4g"]
        template_cols = self._get_template_columns(ws)

        radio4g = self.input_data.get("Radio 4G", pd.DataFrame())
        if radio4g.empty:
            return

        # Get config mappings
        spu_config = self.config.get("SPU", {}).get(self.version, {})
        bandwidth_mapping = spu_config.get("bandwidth_mapping", {})
        band_indicator_mapping = spu_config.get("bandIndicator_mapping", {})
        earfcn_mapping = spu_config.get("earfcn_mapping", {})
        earfcn_ul_mapping = spu_config.get("earfcnUl_mapping", {})

        # Get fixed values from config (priority) and Mapping sheet
        cell4g_config_fixed = spu_config.get("cell4g_fixed_values", {})
        cell4g_fixed_values = dict(cell4g_config_fixed)  # Start with config values

        # Overlay with Mapping sheet values (Mapping sheet takes priority)
        if "cell4g" in self.mappings:
            for mapping in self.mappings["cell4g"]:
                if mapping["fixed_value"] is not None:
                    cell4g_fixed_values[mapping["column"]] = mapping["fixed_value"]

        start_row = self._find_data_start_row(ws)
        row = start_row

        for _, source_row in radio4g.iterrows():
            # Extract source values
            ne_name = safe_str(source_row.get("NE_Name", ""))
            enb_id = safe_str(source_row.get("eNBId", ""))
            cell_id = safe_str(source_row.get("cellId", ""))
            cell_name = safe_str(source_row.get("CellName", ""))
            cell_type = safe_str(source_row.get("CellType", ""))  # FDD or TDD
            mimo = safe_str(source_row.get("MIMO", ""))
            pci = source_row.get("PCI", "")
            rsi = source_row.get("RSI", "")
            cp_spe_ref_sig_pwr = source_row.get("cpSpeRefSigPwr", "")
            tac = source_row.get("TAC", "")
            arfcn_dl = source_row.get("arfcndl", "")
            arfcn_ul = source_row.get("arfcnul", "")
            dl_channel_bandwidth = source_row.get("dlChannelBandwidth", "")
            ul_channel_bandwidth = source_row.get("ulChannelBandwidth", dl_channel_bandwidth)
            rru_port = safe_str(source_row.get("rruPort", ""))
            ri_port_baseband = safe_str(source_row.get("RiPort Baseband", ""))

            # === Direct mappings ===
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=ne_name)
            if "functionUserLabel" in template_cols:
                ws.cell(row=row, column=template_cols["functionUserLabel"], value=ne_name)
            if "eNBId" in template_cols:
                ws.cell(row=row, column=template_cols["eNBId"], value=enb_id)
            if "cellLocalId" in template_cols:
                ws.cell(row=row, column=template_cols["cellLocalId"], value=cell_id)
            if "userLabel" in template_cols:
                ws.cell(row=row, column=template_cols["userLabel"], value=cell_name)
            if "radioMode" in template_cols:
                ws.cell(row=row, column=template_cols["radioMode"], value=cell_type)
            if "pci" in template_cols:
                ws.cell(row=row, column=template_cols["pci"], value=pci)
            if "tac" in template_cols:
                ws.cell(row=row, column=template_cols["tac"], value=tac)
            if "rootSequenceIndex" in template_cols:
                ws.cell(row=row, column=template_cols["rootSequenceIndex"], value=rsi)
            if "cpSpeRefSigPwr" in template_cols:
                ws.cell(row=row, column=template_cols["cpSpeRefSigPwr"], value=cp_spe_ref_sig_pwr)
            if "sectorFunctionId" in template_cols:
                ws.cell(row=row, column=template_cols["sectorFunctionId"], value=cell_name)
            if "cpId" in template_cols:
                ws.cell(row=row, column=template_cols["cpId"], value=cell_id)

            # === Calculated fields ===

            # functionMoId = plmn_eNBId (e.g., 452-04_60486)
            if "functionMoId" in template_cols:
                function_mo_id = f"452-04_{enb_id}"
                ws.cell(row=row, column=template_cols["functionMoId"], value=function_mo_id)

            # bandIndicator: lookup from config based on arfcndl
            if "bandIndicator" in template_cols:
                arfcn_dl_str = str(safe_int(arfcn_dl, 0))
                band_indicator = band_indicator_mapping.get(arfcn_dl_str, "")
                ws.cell(row=row, column=template_cols["bandIndicator"], value=band_indicator)

            # earfcnDl and earfcnUl: only for FDD cells
            # earfcn: only for TDD cells
            if cell_type.upper() == "FDD":
                # earfcnDl: mapping between arfcndl and earfcn_mapping in config.json
                # e.g., if arfcndl is 1700 then earfcnDl is 1855
                if "earfcnDl" in template_cols:
                    arfcn_dl_str = str(safe_int(arfcn_dl, 0))
                    earfcn_dl_mapped = earfcn_mapping.get(arfcn_dl_str, arfcn_dl)
                    ws.cell(row=row, column=template_cols["earfcnDl"], value=earfcn_dl_mapped)

                # earfcnUl: mapping between arfcnul and earfcn_mapping in config.json
                if "earfcnUl" in template_cols:
                    arfcn_ul_str = str(safe_int(arfcn_ul, 0))
                    earfcn_ul_mapped = earfcn_mapping.get(arfcn_ul_str, arfcn_ul)
                    ws.cell(row=row, column=template_cols["earfcnUl"], value=earfcn_ul_mapped)
            else:  # TDD
                # earfcn: mapping between arfcndl and earfcn_mapping in config.json
                if "earfcn" in template_cols:
                    arfcn_dl_str = str(safe_int(arfcn_dl, 0))
                    earfcn_mapped = earfcn_mapping.get(arfcn_dl_str, arfcn_dl)
                    ws.cell(row=row, column=template_cols["earfcn"], value=earfcn_mapped)

            # bandWidthDl/bandWidthUl: map bandwidth (20->5, 15->4, 10->3)
            dl_bw = safe_int(dl_channel_bandwidth, 0)
            ul_bw = safe_int(ul_channel_bandwidth, dl_bw)

            # Calculate bandwidth code: 20->5, 15->4, 10->3, 5->2, 3->1
            def get_bandwidth_code(bw):
                bw_str = str(float(bw)) if bw else "0"
                if bw_str in bandwidth_mapping:
                    return bandwidth_mapping[bw_str]
                # Fallback mapping
                if bw >= 20:
                    return 5
                elif bw >= 15:
                    return 4
                elif bw >= 10:
                    return 3
                elif bw >= 5:
                    return 2
                elif bw >= 3:
                    return 1
                return 0

            if cell_type.upper() == "FDD":
                if "bandWidthDl" in template_cols:
                    ws.cell(row=row, column=template_cols["bandWidthDl"], value=get_bandwidth_code(dl_bw))
                if "bandWidthUl" in template_cols:
                    ws.cell(row=row, column=template_cols["bandWidthUl"], value=get_bandwidth_code(ul_bw))
            else:  # TDD
                if "bandWidth" in template_cols:
                    ws.cell(row=row, column=template_cols["bandWidth"], value=get_bandwidth_code(dl_bw))

            # paForDTCH: 4 if FDD, 2 if TDD
            if "paForDTCH" in template_cols:
                pa_for_dtch = 4 if cell_type.upper() == "FDD" else 2
                ws.cell(row=row, column=template_cols["paForDTCH"], value=pa_for_dtch)

            # cellRSPortNum: MIMO mapping (2T2R->1, 4T4R->2)
            if "cellRSPortNum" in template_cols:
                cell_rs_port_num = 1  # default
                mimo_upper = mimo.upper()
                if "4T4R" in mimo_upper or "4T" in mimo_upper:
                    cell_rs_port_num = 2
                elif "2T2R" in mimo_upper or "2T" in mimo_upper:
                    cell_rs_port_num = 1
                elif "8T8R" in mimo_upper:
                    cell_rs_port_num = 3
                ws.cell(row=row, column=template_cols["cellRSPortNum"], value=cell_rs_port_num)

            # sampleRateCfg: 2 if bandWidthDl >= 4, else 0
            if "sampleRateCfg" in template_cols:
                bw_code = get_bandwidth_code(dl_bw)
                sample_rate_cfg = 2 if bw_code >= 4 else 0
                ws.cell(row=row, column=template_cols["sampleRateCfg"], value=sample_rate_cfg)

            # bpPoolFunctionId: extract from RiPort Baseband (VBP_1_6&OF1 -> BF_6)
            if "bpPoolFunctionId" in template_cols:
                bp_pool_func_id = ""
                if ri_port_baseband:
                    match = re.search(r'VBP_1_(\d+)&', ri_port_baseband)
                    if match:
                        bp_pool_func_id = f"BF_{match.group(1)}"
                ws.cell(row=row, column=template_cols["bpPoolFunctionId"], value=bp_pool_func_id)

            # anttoPortMap: based on rruPort and CellType
            if "anttoPortMap" in template_cols:
                ant_to_port_map = self._get_ant_to_port_map(rru_port, cell_type)
                ws.cell(row=row, column=template_cols["anttoPortMap"], value=ant_to_port_map)

            # nRearfcn: Get SSB frequency from Radio 5G based on Relation 5G column
            relation_5g = safe_str(source_row.get("Relation 5G", ""))
            if "nRearfcn" in template_cols and relation_5g:
                radio5g = self.input_data.get("Radio 5G", pd.DataFrame())
                if not radio5g.empty and "ssbFrequency" in radio5g.columns:
                    # Find matching 5G cell by name
                    matching_5g = radio5g[radio5g["nRCell"] == relation_5g]
                    if not matching_5g.empty:
                        ssb_freq = matching_5g.iloc[0].get("ssbFrequency", "")
                        if pd.notna(ssb_freq):
                            ws.cell(row=row, column=template_cols["nRearfcn"], value=ssb_freq)

            # Apply fixed values from config and Mapping sheet
            for col_name, fixed_val in cell4g_fixed_values.items():
                if col_name in template_cols:
                    # Skip fields that have been calculated or directly mapped
                    already_set = [
                        "meId", "functionUserLabel", "eNBId", "cellLocalId", "userLabel",
                        "radioMode", "pci", "tac", "rootSequenceIndex", "cpSpeRefSigPwr",
                        "sectorFunctionId", "earfcnDl", "cpId", "functionMoId", "bandIndicator",
                        "earfcnUl", "earfcn", "bandWidthDl", "bandWidthUl", "bandWidth", "paForDTCH",
                        "cellRSPortNum", "sampleRateCfg", "bpPoolFunctionId", "anttoPortMap",
                        "nRearfcn"
                    ]
                    if col_name not in already_set:
                        ws.cell(row=row, column=template_cols[col_name], value=fixed_val)

            row += 1

    def _get_ant_to_port_map(self, rru_port, cell_type):
        """Get anttoPortMap based on rruPort and CellType.

        Mapping rules:
        - rruPort 12: '0;1;15;15;...' (64 values, first 2 are 0;1, rest are 15)
        - rruPort 34: '15;15;0;1;15;15;...' (64 values, positions 3,4 are 0;1)
        - rruPort 1234: '0;1;2;3;15;15;...' (64 values, first 4 are 0;1;2;3)
        - TDD mode: '0;2;0;2;0;2;0;2;0;2;0;2;0;2;0;2;1;3;1;3;...' (64 values, pattern)
        """
        # For TDD, use special pattern
        if cell_type.upper() == "TDD":
            # TDD pattern: 0;2 repeated 16 times, then 1;3 repeated 16 times, then repeat
            tdd_pattern = []
            for _ in range(8):
                tdd_pattern.extend([0, 2])
            for _ in range(8):
                tdd_pattern.extend([1, 3])
            for _ in range(8):
                tdd_pattern.extend([0, 2])
            for _ in range(8):
                tdd_pattern.extend([1, 3])
            return ";".join(str(x) for x in tdd_pattern)

        # For FDD, map based on rruPort
        base_pattern = [15] * 64  # Default: all 15

        rru_port_str = str(rru_port).strip()

        if rru_port_str == "12":
            # Ports 1-2: positions 0,1
            base_pattern[0] = 0
            base_pattern[1] = 1
        elif rru_port_str == "34":
            # Ports 3-4: positions 2,3
            base_pattern[2] = 0
            base_pattern[3] = 1
        elif rru_port_str == "1234":
            # Ports 1-4: positions 0,1,2,3
            base_pattern[0] = 0
            base_pattern[1] = 1
            base_pattern[2] = 2
            base_pattern[3] = 3
        elif rru_port_str == "1":
            base_pattern[0] = 0
        elif rru_port_str == "2":
            base_pattern[1] = 1
        elif rru_port_str == "3":
            base_pattern[2] = 0
        elif rru_port_str == "4":
            base_pattern[3] = 1

        return ";".join(str(x) for x in base_pattern)

    def _process_ip_sheet(self, wb):
        """Process Ip sheet.

        Mapping rules:
        - meId: NE_Name from IP sheet (Input folder)
        - moId: OAM, LTE, or NR

        For OAM:
        - ipAddress: OAM_IP from IP sheet
        - prefixLength: fixed "30"
        - gatewayIp: OAM_Gateway from IP sheet
        - vid: OAM_vlan from IP sheet

        For LTE:
        - ipAddress: LTE_IP from IP sheet
        - prefixLength: fixed "30"
        - gatewayIp: LTE_Gateway from IP sheet
        - vid: LTE_vlan from IP sheet
        - serviceMapRadioType: fixed "LTE"
        - serviceInterfaceType: fixed "4;8;32"
        - plmn: fixed "452-04"

        For NR (if NR_IP exists):
        - ipAddress: NR_IP from IP sheet
        - prefixLength: fixed "30"
        - gatewayIp: NR_Gateway from IP sheet
        - vid: NR_vlan from IP sheet
        - serviceMapRadioType: fixed "5G"
        - serviceInterfaceType: fixed "1;2;4;16;32"
        - plmn: fixed "452-04"
        """
        if "Ip" not in wb.sheetnames:
            return

        ws = wb["Ip"]
        template_cols = self._get_template_columns(ws)
        ip_df = self.input_data.get("IP", pd.DataFrame())

        if ip_df.empty:
            return

        row = self._find_data_start_row(ws)
        for _, ip_row in ip_df.iterrows():
            ne_name = ip_row.get("NE_Name", "")

            # OAM IP
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=ne_name)
            if "moId" in template_cols:
                ws.cell(row=row, column=template_cols["moId"], value="OAM")
            if "ipAddress" in template_cols:
                ws.cell(row=row, column=template_cols["ipAddress"], value=ip_row.get("OAM_IP", ""))
            if "prefixLength" in template_cols:
                ws.cell(row=row, column=template_cols["prefixLength"], value="30")
            if "gatewayIp" in template_cols:
                ws.cell(row=row, column=template_cols["gatewayIp"], value=ip_row.get("OAM_Gateway", ""))
            if "vid" in template_cols:
                oam_vlan = ip_row.get("OAM_vlan", "")
                if pd.notna(oam_vlan):
                    ws.cell(row=row, column=template_cols["vid"], value=oam_vlan)
            row += 1

            # LTE IP
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=ne_name)
            if "moId" in template_cols:
                ws.cell(row=row, column=template_cols["moId"], value="LTE")
            if "ipAddress" in template_cols:
                ws.cell(row=row, column=template_cols["ipAddress"], value=ip_row.get("LTE_IP", ""))
            if "prefixLength" in template_cols:
                ws.cell(row=row, column=template_cols["prefixLength"], value="30")
            if "gatewayIp" in template_cols:
                ws.cell(row=row, column=template_cols["gatewayIp"], value=ip_row.get("LTE_Gateway", ""))
            if "vid" in template_cols:
                lte_vlan = ip_row.get("LTE_vlan", "")
                if pd.notna(lte_vlan):
                    ws.cell(row=row, column=template_cols["vid"], value=lte_vlan)
            if "serviceMapRadioType" in template_cols:
                ws.cell(row=row, column=template_cols["serviceMapRadioType"], value="LTE")
            if "serviceInterfaceType" in template_cols:
                ws.cell(row=row, column=template_cols["serviceInterfaceType"], value="4;8;32")
            if "plmn" in template_cols:
                ws.cell(row=row, column=template_cols["plmn"], value="452-04")
            row += 1

            # NR IP (only if NR_IP exists)
            if pd.notna(ip_row.get("NR_IP")):
                if "meId" in template_cols:
                    ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                if "moId" in template_cols:
                    ws.cell(row=row, column=template_cols["moId"], value="NR")
                if "ipAddress" in template_cols:
                    ws.cell(row=row, column=template_cols["ipAddress"], value=ip_row.get("NR_IP", ""))
                if "prefixLength" in template_cols:
                    ws.cell(row=row, column=template_cols["prefixLength"], value="30")
                if "gatewayIp" in template_cols:
                    ws.cell(row=row, column=template_cols["gatewayIp"], value=ip_row.get("NR_Gateway", ""))
                if "vid" in template_cols:
                    nr_vlan = ip_row.get("NR_vlan", "")
                    if pd.notna(nr_vlan):
                        ws.cell(row=row, column=template_cols["vid"], value=nr_vlan)
                if "serviceMapRadioType" in template_cols:
                    ws.cell(row=row, column=template_cols["serviceMapRadioType"], value="5G")
                if "serviceInterfaceType" in template_cols:
                    ws.cell(row=row, column=template_cols["serviceInterfaceType"], value="1;2;4;16;32")
                if "plmn" in template_cols:
                    ws.cell(row=row, column=template_cols["plmn"], value="452-04")
                row += 1

    def _process_sctp_sheet(self, wb):
        """Process Sctp sheet.

        SCTP numbering and configuration rules:
        - MME entries (sctpNo 1 to N, where N = number of MME values in IP sheet):
          localPort=36412, remotePort=36412, radioMode=48, assoType=1, NBId=eNBId
          localIp=LTE_IP (repeated for each remoteIp), remoteIp from config.json mme mapping

        - AMF entries (sctpNo N+1 to N+M, where M = number of AMF values in IP sheet):
          localPort=38412, remotePort=38412, radioMode=8192, assoType=1, NBId=gNBId
          localIp=NR_IP (repeated for each remoteIp), remoteIp from config.json amf mapping

        - X2 interface entries (next 2 sctpNo values after AMF):
          First X2: localPort=36422, remotePort=36422, assoType=5, radioMode=48,
                   localIp=LTE_IP, remoteIp=NR_IP, NBId=eNBId
          Second X2: localPort=36422, remotePort=36422, assoType=5, radioMode=8192,
                    localIp=NR_IP, remoteIp=LTE_IP, NBId=gNBId

        - pLMNId = 452-04 for all entries
        """
        if "Sctp" not in wb.sheetnames:
            return

        ws = wb["Sctp"]
        template_cols = self._get_template_columns(ws)
        ip_df = self.input_data.get("IP", pd.DataFrame())

        if ip_df.empty:
            return

        mme_config = self.config.get("mme", {})
        amf_config = self.config.get("amf", {})

        row = self._find_data_start_row(ws)

        for _, ip_row in ip_df.iterrows():
            ne_name = safe_str(ip_row.get("NE_Name", ""))
            lte_ip = safe_str(ip_row.get("LTE_IP", ""))
            nr_ip = safe_str(ip_row.get("NR_IP", ""))
            enb_id = safe_str(ip_row.get("eNBId", ""))
            gnb_id = safe_str(ip_row.get("gNBId", ""))

            sctp_no = 1

            # Process MME entries (port 36412, radioMode 48, assoType 1)
            mme_str = safe_str(ip_row.get("MME", ""))
            if mme_str:
                for mme_name in mme_str.split():
                    if mme_name in mme_config:
                        # Get all remote IPs for this MME (excluding 0.0.0.0)
                        remote_ips = [ip for ip in mme_config[mme_name] if ip and ip != "0.0.0.0"]
                        if remote_ips:
                            # Build remoteIp string with semicolons
                            remote_ip_str = ";".join(remote_ips)
                            # Build localIp string - repeat LTE_IP for each remote IP
                            local_ip_str = ";".join([lte_ip] * len(remote_ips))

                            # Write MME SCTP entry
                            if "meId" in template_cols:
                                ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                            if "sctpNo" in template_cols:
                                ws.cell(row=row, column=template_cols["sctpNo"], value=sctp_no)
                            if "localPort" in template_cols:
                                ws.cell(row=row, column=template_cols["localPort"], value=36412)
                            if "localIp" in template_cols:
                                ws.cell(row=row, column=template_cols["localIp"], value=local_ip_str)
                            if "remotePort" in template_cols:
                                ws.cell(row=row, column=template_cols["remotePort"], value=36412)
                            if "remoteIp" in template_cols:
                                ws.cell(row=row, column=template_cols["remoteIp"], value=remote_ip_str)
                            if "radioMode" in template_cols:
                                ws.cell(row=row, column=template_cols["radioMode"], value=48)
                            if "assoType" in template_cols:
                                ws.cell(row=row, column=template_cols["assoType"], value=1)
                            if "NBId" in template_cols:
                                ws.cell(row=row, column=template_cols["NBId"], value=enb_id)
                            if "pLMNId" in template_cols:
                                ws.cell(row=row, column=template_cols["pLMNId"], value="452-04")

                            sctp_no += 1
                            row += 1

            # Process AMF entries (port 38412, radioMode 8192, assoType 1)
            amf_str = safe_str(ip_row.get("AMF", ""))
            if amf_str and nr_ip:
                for amf_name in amf_str.split():
                    if amf_name in amf_config:
                        # Get all remote IPs for this AMF (excluding 0.0.0.0)
                        remote_ips = [ip for ip in amf_config[amf_name] if ip and ip != "0.0.0.0"]
                        if remote_ips:
                            # Build remoteIp string with semicolons
                            remote_ip_str = ";".join(remote_ips)
                            # Build localIp string - repeat NR_IP for each remote IP
                            local_ip_str = ";".join([nr_ip] * len(remote_ips))

                            # Write AMF SCTP entry
                            if "meId" in template_cols:
                                ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                            if "sctpNo" in template_cols:
                                ws.cell(row=row, column=template_cols["sctpNo"], value=sctp_no)
                            if "localPort" in template_cols:
                                ws.cell(row=row, column=template_cols["localPort"], value=38412)
                            if "localIp" in template_cols:
                                ws.cell(row=row, column=template_cols["localIp"], value=local_ip_str)
                            if "remotePort" in template_cols:
                                ws.cell(row=row, column=template_cols["remotePort"], value=38412)
                            if "remoteIp" in template_cols:
                                ws.cell(row=row, column=template_cols["remoteIp"], value=remote_ip_str)
                            if "radioMode" in template_cols:
                                ws.cell(row=row, column=template_cols["radioMode"], value=8192)
                            if "assoType" in template_cols:
                                ws.cell(row=row, column=template_cols["assoType"], value=1)
                            if "NBId" in template_cols:
                                ws.cell(row=row, column=template_cols["NBId"], value=gnb_id)
                            if "pLMNId" in template_cols:
                                ws.cell(row=row, column=template_cols["pLMNId"], value="452-04")

                            sctp_no += 1
                            row += 1

            # Process X2 interface entries (port 36422, assoType 5)
            # Only add X2 entries if both LTE_IP and NR_IP are available
            if lte_ip and nr_ip:
                # First X2 entry: LTE -> NR (radioMode 48, NBId = eNBId)
                if "meId" in template_cols:
                    ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                if "sctpNo" in template_cols:
                    ws.cell(row=row, column=template_cols["sctpNo"], value=sctp_no)
                if "localPort" in template_cols:
                    ws.cell(row=row, column=template_cols["localPort"], value=36422)
                if "localIp" in template_cols:
                    ws.cell(row=row, column=template_cols["localIp"], value=lte_ip)
                if "remotePort" in template_cols:
                    ws.cell(row=row, column=template_cols["remotePort"], value=36422)
                if "remoteIp" in template_cols:
                    ws.cell(row=row, column=template_cols["remoteIp"], value=nr_ip)
                if "radioMode" in template_cols:
                    ws.cell(row=row, column=template_cols["radioMode"], value=48)
                if "assoType" in template_cols:
                    ws.cell(row=row, column=template_cols["assoType"], value=5)
                if "NBId" in template_cols:
                    ws.cell(row=row, column=template_cols["NBId"], value=enb_id)
                if "pLMNId" in template_cols:
                    ws.cell(row=row, column=template_cols["pLMNId"], value="452-04")

                sctp_no += 1
                row += 1

                # Second X2 entry: NR -> LTE (radioMode 8192, NBId = gNBId)
                if "meId" in template_cols:
                    ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                if "sctpNo" in template_cols:
                    ws.cell(row=row, column=template_cols["sctpNo"], value=sctp_no)
                if "localPort" in template_cols:
                    ws.cell(row=row, column=template_cols["localPort"], value=36422)
                if "localIp" in template_cols:
                    ws.cell(row=row, column=template_cols["localIp"], value=nr_ip)
                if "remotePort" in template_cols:
                    ws.cell(row=row, column=template_cols["remotePort"], value=36422)
                if "remoteIp" in template_cols:
                    ws.cell(row=row, column=template_cols["remoteIp"], value=lte_ip)
                if "radioMode" in template_cols:
                    ws.cell(row=row, column=template_cols["radioMode"], value=8192)
                if "assoType" in template_cols:
                    ws.cell(row=row, column=template_cols["assoType"], value=5)
                if "NBId" in template_cols:
                    ws.cell(row=row, column=template_cols["NBId"], value=gnb_id)
                if "pLMNId" in template_cols:
                    ws.cell(row=row, column=template_cols["pLMNId"], value="452-04")

                sctp_no += 1
                row += 1

    def _process_bbu_sheet(self, wb):
        """Process BBU sheet."""
        if "BBU" not in wb.sheetnames:
            return

        ws = wb["BBU"]
        template_cols = self._get_template_columns(ws)
        ip_df = self.input_data.get("IP", pd.DataFrame())

        if ip_df.empty:
            return

        bb_configs = self.config.get("SPU", {}).get(self.version, {}).get("baseband_configs", {})

        row = self._find_data_start_row(ws)
        for _, ip_row in ip_df.iterrows():
            ne_name = ip_row.get("NE_Name", "")
            baseband = safe_str(ip_row.get("Baseband config", ""))

            if baseband in bb_configs:
                for bb in bb_configs[baseband]:
                    if "meId" in template_cols:
                        ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                    if "moId" in template_cols:
                        ws.cell(row=row, column=template_cols["moId"], value=bb.get("moId", ""))
                    if "name" in template_cols:
                        ws.cell(row=row, column=template_cols["name"], value=bb.get("name", ""))
                    if "hwWorkScence" in template_cols:
                        ws.cell(row=row, column=template_cols["hwWorkScence"], value=bb.get("hwWorkScence", ""))
                    if "functionMode" in template_cols:
                        ws.cell(row=row, column=template_cols["functionMode"], value=bb.get("functionMode", ""))
                    if "bpPoolFunction" in template_cols:
                        ws.cell(row=row, column=template_cols["bpPoolFunction"], value=bb.get("bpPoolFunction", ""))
                    if "vswPortInfo" in template_cols:
                        ws.cell(row=row, column=template_cols["vswPortInfo"], value=bb.get("vswPortInfo", ""))
                    if "interconnectionPortInfo" in template_cols:
                        ws.cell(row=row, column=template_cols["interconnectionPortInfo"], value=bb.get("interconnectionPortInfo", ""))
                    if "protocolType" in template_cols:
                        ws.cell(row=row, column=template_cols["protocolType"], value=bb.get("protocolType", ""))
                    if "bitRateOnIrLine" in template_cols:
                        ws.cell(row=row, column=template_cols["bitRateOnIrLine"], value=bb.get("bitRateOnIrLine", ""))
                    if "peerInterconnectionPortInfo" in template_cols:
                        ws.cell(row=row, column=template_cols["peerInterconnectionPortInfo"], value=bb.get("peerInterconnectionPortInfo", ""))
                    if "refLinkRu" in template_cols:
                        ws.cell(row=row, column=template_cols["refLinkRu"], value=bb.get("refLinkRu", ""))
                    if "refCabinet" in template_cols:
                        ws.cell(row=row, column=template_cols["refCabinet"], value=bb.get("refCabinet", ""))
                    if "mainCtrlWorkMode" in template_cols:
                        ws.cell(row=row, column=template_cols["mainCtrlWorkMode"], value=bb.get("mainCtrlWorkMode", ""))
                    if "networkBackupMode" in template_cols:
                        ws.cell(row=row, column=template_cols["networkBackupMode"], value=bb.get("networkBackupMode", ""))
                    if "refReplaceableUnit" in template_cols:
                        ws.cell(row=row, column=template_cols["refReplaceableUnit"], value=bb.get("refReplaceableUnit", ""))
                    if "redundancyModeSwitch" in template_cols:
                        ws.cell(row=row, column=template_cols["redundancyModeSwitch"], value=bb.get("redundancyModeSwitch", ""))
                    if "masterWorkMode" in template_cols:
                        ws.cell(row=row, column=template_cols["masterWorkMode"], value=bb.get("masterWorkMode", ""))
                    if "specOption" in template_cols:
                        ws.cell(row=row, column=template_cols["specOption"], value=bb.get("specOption", ""))
                    row += 1
            else:
                # Placeholder row
                if "meId" in template_cols:
                    ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                row += 1

    def _process_drycontact_sheet(self, wb):
        """Process DryContactCable sheet.

        Mapping rules:
        - meId: NE_Name from IP sheet
        - moId: from drycontact_configs in config.json
        - refDryContactPort: from drycontact_configs in config.json
        - dryContactNo: from drycontact_configs in config.json
        - dryType: from drycontact_configs in config.json
        - alarmStatus: from drycontact_configs in config.json
        - alarmNameOfInput: from drycontact_configs in config.json
        - alarmNameOfOutput: from drycontact_configs in config.json
        """
        if "DryContactCable" not in wb.sheetnames:
            return

        ws = wb["DryContactCable"]
        template_cols = self._get_template_columns(ws)
        ip_df = self.input_data.get("IP", pd.DataFrame())

        if ip_df.empty:
            return

        dc_configs = self.config.get("SPU", {}).get(self.version, {}).get("drycontact_configs", [])

        row = self._find_data_start_row(ws)
        for _, ip_row in ip_df.iterrows():
            ne_name = ip_row.get("NE_Name", "")

            for dc in dc_configs:
                if "meId" in template_cols:
                    ws.cell(row=row, column=template_cols["meId"], value=ne_name)
                if "moId" in template_cols:
                    ws.cell(row=row, column=template_cols["moId"], value=dc.get("moId", ""))
                if "refDryContactPort" in template_cols:
                    ws.cell(row=row, column=template_cols["refDryContactPort"], value=dc.get("refDryContactPort", ""))
                if "dryContactNo" in template_cols:
                    ws.cell(row=row, column=template_cols["dryContactNo"], value=dc.get("dryContactNo", ""))
                if "dryType" in template_cols:
                    ws.cell(row=row, column=template_cols["dryType"], value=dc.get("dryType", ""))
                if "alarmStatus" in template_cols:
                    ws.cell(row=row, column=template_cols["alarmStatus"], value=dc.get("alarmStatus", ""))
                if "alarmNameOfInput" in template_cols:
                    ws.cell(row=row, column=template_cols["alarmNameOfInput"], value=dc.get("alarmNameOfInput", ""))
                if "alarmNameOfOutput" in template_cols:
                    ws.cell(row=row, column=template_cols["alarmNameOfOutput"], value=dc.get("alarmNameOfOutput", ""))
                row += 1

    def _process_cable_sheet(self, wb):
        """Process cable sheet with RiPort parsing.

        Mapping rules:
        - meId: NE_Name from Radio 4G/5G (same as IP sheet)
        - upRiDevice: Before '&' of 'RiPort Baseband' (e.g., 'VBP_1_6')
        - upRiPort: After '&' of 'RiPort Baseband' (e.g., 'OF1')
        - downRiDevice: RRU column from Radio 4G/5G (e.g., 'RRU_1', 'AAU_1')
        - downRiPort: 'RiPort RRU' column (e.g., 'OPT1')
        - upBitRateOnIrLine: '25#0' if AAU, '11#0' if RRU
        - downBitRateOnIrLine: '25#0' if AAU, '255#0' if RRU
        - upProtocolType: 7 if AAU, 0 if RRU
        - downProtocolType: 7 if AAU, 0 if RRU
        """
        if "cable" not in wb.sheetnames:
            return

        ws = wb["cable"]
        template_cols = self._get_template_columns(ws)

        radio4g = self.input_data.get("Radio 4G", pd.DataFrame())
        radio5g = self.input_data.get("Radio 5G", pd.DataFrame())

        # Collect all cable entries (unique per NE_Name + RRU combination)
        # Key: (ne_name, rru_name), Value: cable info dict
        cable_entries = {}

        # Process 4G Radio entries
        if not radio4g.empty and "RRU" in radio4g.columns:
            for _, row_data in radio4g.iterrows():
                ne_name = safe_str(row_data.get("NE_Name", ""))
                rru_name = safe_str(row_data.get("RRU", ""))
                ri_baseband = safe_str(row_data.get("RiPort Baseband", ""))
                ri_rru = safe_str(row_data.get("RiPort RRU", ""))

                if not ne_name or not rru_name or not ri_baseband:
                    continue

                key = (ne_name, rru_name)
                if key in cable_entries:
                    continue  # Already processed this RRU

                # Parse RiPort Baseband: "VBP_1_6&OF1" -> upRiDevice="VBP_1_6", upRiPort="OF1"
                if "&" in ri_baseband:
                    parts = ri_baseband.split("&", 1)
                    up_ri_device = parts[0].strip()
                    up_ri_port = parts[1].strip() if len(parts) > 1 else ""
                else:
                    up_ri_device = ri_baseband
                    up_ri_port = ""

                # Determine if AAU or RRU based on downRiDevice (RRU column)
                # AAU devices typically start with "AAU", RRU devices start with "RRU"
                is_aau = rru_name.upper().startswith("AAU")

                cable_entries[key] = {
                    "ne_name": ne_name,
                    "up_ri_device": up_ri_device,
                    "up_ri_port": up_ri_port,
                    "down_ri_device": rru_name,
                    "down_ri_port": ri_rru,
                    "is_aau": is_aau
                }

        # Process 5G Radio entries
        if not radio5g.empty and "RRU" in radio5g.columns:
            for _, row_data in radio5g.iterrows():
                ne_name = safe_str(row_data.get("NE_Name", ""))
                rru_name = safe_str(row_data.get("RRU", ""))
                ri_baseband = safe_str(row_data.get("RiPort Baseband", ""))
                ri_rru = safe_str(row_data.get("RiPort RRU", ""))

                if not ne_name or not rru_name or not ri_baseband:
                    continue

                key = (ne_name, rru_name)
                if key in cable_entries:
                    continue  # Already processed this RRU

                # Parse RiPort Baseband
                if "&" in ri_baseband:
                    parts = ri_baseband.split("&", 1)
                    up_ri_device = parts[0].strip()
                    up_ri_port = parts[1].strip() if len(parts) > 1 else ""
                else:
                    up_ri_device = ri_baseband
                    up_ri_port = ""

                # 5G typically uses AAU
                is_aau = rru_name.upper().startswith("AAU")

                cable_entries[key] = {
                    "ne_name": ne_name,
                    "up_ri_device": up_ri_device,
                    "up_ri_port": up_ri_port,
                    "down_ri_device": rru_name,
                    "down_ri_port": ri_rru,
                    "is_aau": is_aau
                }

        # Write cable entries to worksheet
        start_row = self._find_data_start_row(ws)
        row = start_row

        for (ne_name, rru_name), entry in cable_entries.items():
            is_aau = entry["is_aau"]

            # Set bit rate and protocol based on device type
            if is_aau:
                up_bit_rate = "25#0"
                down_bit_rate = "25#0"
                up_protocol = 7
                down_protocol = 7
            else:
                up_bit_rate = "11#0"
                down_bit_rate = "255#0"
                up_protocol = 0
                down_protocol = 0

            # Write to worksheet
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=entry["ne_name"])
            if "upRiDevice" in template_cols:
                ws.cell(row=row, column=template_cols["upRiDevice"], value=entry["up_ri_device"])
            if "upRiPort" in template_cols:
                ws.cell(row=row, column=template_cols["upRiPort"], value=entry["up_ri_port"])
            if "downRiDevice" in template_cols:
                ws.cell(row=row, column=template_cols["downRiDevice"], value=entry["down_ri_device"])
            if "downRiPort" in template_cols:
                ws.cell(row=row, column=template_cols["downRiPort"], value=entry["down_ri_port"])
            if "upBitRateOnIrLine" in template_cols:
                ws.cell(row=row, column=template_cols["upBitRateOnIrLine"], value=up_bit_rate)
            if "downBitRateOnIrLine" in template_cols:
                ws.cell(row=row, column=template_cols["downBitRateOnIrLine"], value=down_bit_rate)
            if "upProtocolType" in template_cols:
                ws.cell(row=row, column=template_cols["upProtocolType"], value=up_protocol)
            if "downProtocolType" in template_cols:
                ws.cell(row=row, column=template_cols["downProtocolType"], value=down_protocol)

            row += 1

    def _process_aisgctrlport_sheet(self, wb):
        """Process AisgCtrlPort sheet.

        Mapping rules:
        - meId: NE_Name from IP sheet
        - ruId: RRU from Radio 4G and Radio 5G sheets (unique per NE)
        - moId: fixed value "AISG"
        - powerSupplySwitch: fixed value "1"
        - outputVoltageAISG: fixed value "22"
        """
        if "AisgCtrlPort" not in wb.sheetnames:
            return

        ws = wb["AisgCtrlPort"]
        template_cols = self._get_template_columns(ws)

        radio4g = self.input_data.get("Radio 4G", pd.DataFrame())
        radio5g = self.input_data.get("Radio 5G", pd.DataFrame())

        # Collect unique (ne_name, rru) combinations from Radio 4G and Radio 5G
        aisg_entries = {}

        # Process Radio 4G
        if not radio4g.empty and "RRU" in radio4g.columns and "NE_Name" in radio4g.columns:
            for _, row_data in radio4g.iterrows():
                ne_name = safe_str(row_data.get("NE_Name", ""))
                rru = safe_str(row_data.get("RRU", ""))

                if not ne_name or not rru:
                    continue

                key = (ne_name, rru)
                if key not in aisg_entries:
                    aisg_entries[key] = {"ne_name": ne_name, "rru": rru}

        # Process Radio 5G
        if not radio5g.empty and "RRU" in radio5g.columns and "NE_Name" in radio5g.columns:
            for _, row_data in radio5g.iterrows():
                ne_name = safe_str(row_data.get("NE_Name", ""))
                rru = safe_str(row_data.get("RRU", ""))

                if not ne_name or not rru:
                    continue

                key = (ne_name, rru)
                if key not in aisg_entries:
                    aisg_entries[key] = {"ne_name": ne_name, "rru": rru}

        # Write to worksheet
        start_row = self._find_data_start_row(ws)
        row = start_row

        for (ne_name, rru), entry in aisg_entries.items():
            if "meId" in template_cols:
                ws.cell(row=row, column=template_cols["meId"], value=entry["ne_name"])
            if "ruId" in template_cols:
                ws.cell(row=row, column=template_cols["ruId"], value=entry["rru"])
            if "moId" in template_cols:
                ws.cell(row=row, column=template_cols["moId"], value="AISG")
            if "powerSupplySwitch" in template_cols:
                ws.cell(row=row, column=template_cols["powerSupplySwitch"], value="1")
            if "outputVoltageAISG" in template_cols:
                ws.cell(row=row, column=template_cols["outputVoltageAISG"], value="22")

            row += 1
